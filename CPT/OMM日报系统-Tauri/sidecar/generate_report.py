import os
import sys
import re
import shutil
import subprocess
import math
import random
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, time as dtime


def format_minutes(minutes):
    """将分钟格式化为 xhxxm 字符串。"""
    if minutes is None:
        return ''
    minutes = int(minutes)
    h = minutes // 60
    m = minutes % 60
    if h > 0:
        return f'{h}h{m:02d}m'
    return f'{m}m'


# ==================== 配置 ====================
# 工作目录：
# - 脚本运行时优先使用环境变量 YX_WORK_DIR 或 GUI 传入的值
# - 否则回退到脚本/EXE 所在目录
if getattr(sys, 'frozen', False):
    _DEFAULT_WORK_DIR = os.path.dirname(sys.executable)
else:
    _DEFAULT_WORK_DIR = os.path.dirname(os.path.abspath(__file__))

WORK_DIR = os.environ.get('YX_WORK_DIR', _DEFAULT_WORK_DIR)
OUTPUT_DIR = os.path.join(WORK_DIR, "TEST")
OPERATOR_NAME = "禹欣"


_TEMPLATE_CANDIDATES = [
    "滁州量测室总体日报汇总表-OMM-禹欣1.xlsx",
    "滁州量测室总体日报汇总表-OMM.xlsx",
    "滁州量测室总体日报汇总表-OMM",
]


def _bundled_template_candidates():
    """推导便携版/开发模式下可能存在的内置模板路径。"""
    bases = []
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        bases.extend([
            exe_dir,
            os.path.dirname(exe_dir),  # 便携版中 sidecar 常位于 binaries/
        ])
    else:
        sidecar_dir = os.path.dirname(os.path.abspath(__file__))
        bases.extend([
            sidecar_dir,
            os.path.dirname(sidecar_dir),
            os.path.join(os.path.dirname(sidecar_dir), 'src-tauri'),
        ])

    seen = set()
    for base in bases:
        if not base or base in seen:
            continue
        seen.add(base)
        yield os.path.join(base, 'resources', 'template.xlsx')
        yield os.path.join(base, 'template.xlsx')


def _find_template(work_dir):
    """在工作目录中查找模板文件。

    查找优先级：
    1. 环境变量 YX_USER_TEMPLATE（用户通过「替换模板」按钮指定的自定义模板）
    2. 工作目录根的候选名精确匹配
    3. 工作目录直接子目录的候选名精确匹配
    4. 兜底：任意含"日报汇总表"的 .xlsx 文件
    5. 软件内置模板：环境变量 YX_BUNDLED_TEMPLATE 或 exe/resources/template.xlsx
    """
    # 0. 用户自定义模板（通过设置界面「替换模板」按钮指定，最高优先级）
    user_tpl = os.environ.get('YX_USER_TEMPLATE')
    if user_tpl and os.path.isfile(user_tpl):
        return user_tpl

    # 1. 工作目录查找（用户放在工作目录的模板仍优先于内置模板）
    if work_dir and os.path.isdir(work_dir):
        # 1a. 精确匹配候选名
        for name in _TEMPLATE_CANDIDATES:
            p = os.path.join(work_dir, name)
            if os.path.isfile(p):
                return p
        # 1b. 所有直接子目录（精确匹配）
        try:
            for sub in os.listdir(work_dir):
                sub_path = os.path.join(work_dir, sub)
                if not os.path.isdir(sub_path):
                    continue
                for name in _TEMPLATE_CANDIDATES:
                    p = os.path.join(sub_path, name)
                    if os.path.isfile(p):
                        return p
        except OSError:
            pass
        # 1c. 兜底：任意含"日报汇总表"的 .xlsx 文件
        try:
            for fname in os.listdir(work_dir):
                fpath = os.path.join(work_dir, fname)
                if os.path.isfile(fpath) and fname.lower().endswith('.xlsx') and '日报汇总表' in fname:
                    return fpath
        except OSError:
            pass

    # 2. 软件内置打包模板（最终兜底，保证无工作目录模板时也能生成）
    bundled_tpl = os.environ.get('YX_BUNDLED_TEMPLATE')
    if bundled_tpl and os.path.isfile(bundled_tpl):
        return bundled_tpl
    for bundled_tpl in _bundled_template_candidates():
        if os.path.isfile(bundled_tpl):
            return bundled_tpl

    return None


TEMPLATE_PATH = _find_template(WORK_DIR)


def refresh_template():
    """重新解析模板路径（用户替换模板后调用）。返回最新模板路径或 None。"""
    global TEMPLATE_PATH
    TEMPLATE_PATH = _find_template(WORK_DIR)
    return TEMPLATE_PATH


def set_work_dir(work_dir):
    """设置/切换工作目录，供 GUI 调用。返回是否成功。"""
    global WORK_DIR, TEMPLATE_PATH, OUTPUT_DIR
    if not work_dir or not os.path.isdir(work_dir):
        return False
    WORK_DIR = os.path.abspath(work_dir)
    TEMPLATE_PATH = _find_template(WORK_DIR)
    OUTPUT_DIR = os.path.join(WORK_DIR, "TEST")
    os.environ['YX_WORK_DIR'] = WORK_DIR
    return True


def get_work_dir():
    return WORK_DIR


def _is_zhengxing_cnc(folder_name):
    """判断原始文件夹名是否同时包含'整形'和'CNC'。"""
    if not folder_name:
        return False
    return '整形' in folder_name and 'CNC' in folder_name

SHIFTS = {
    'A': {
        'breaks': [(120, 130), (230, 290), (430, 440), (520, 560)],
        'start_offset': 480,
        'name': '白班',
        'early_max': 560,  # 17:20 = 560min from 8:00
        'early_breaks': [(120, 130), (230, 290), (430, 440), (520, 560)],
    },
    'B': {
        'breaks': [(120, 130), (200, 260), (430, 440), (580, 620)],
        'start_offset': 1200,
        'name': '夜班',
        'early_max': 560,  # 05:20 = 560min from 20:00
        'early_breaks': [(120, 130), (200, 260), (430, 440)],
    },
}
MAX_SHIFT = 720
TARGET_WORK_NORMAL = 570
TARGET_WORK_EARLY = 450
TARGET_END_NORMAL = 720
MIN_TPP = 1.5
EARLY_MAX = 560
TPP_MIN_DEFAULT = 2.0
TPP_MAX_DEFAULT = 5.0
PKG_REST_DEFAULT = 0


def _field_strong_pattern_detected(field, folder_name, parts):
    """判断文件夹名中是否包含某字段的强识别特征（表示用户很可能写了该字段）。"""
    fn = folder_name
    fnu = folder_name.upper()
    if field == 'station':
        return any(k in fn for k in ('射出', '烧结', '整形', '焊接', 'CNC', '镭雕', 'BIN2', 'BIN4',
                                      'FAI', '测试片', '测试', '孔内直径', '外径', '内径',
                                      '直径', '圆度', 'cam件', 'AOI', '对标')) or \
               re.search(r'\b[A-Z]QC\b', fnu) is not None
    if field == 'product':
        return bool(re.match(r'^(\d{3})\b', folder_name)) or bool(re.search(r'(\d{5})', folder_name))
    if field == 'work_order':
        return 'WW' in fnu or any(k in p for p in parts for k in ('工单号', '工单', '单号'))
    if field == 'mold':
        return 'BIN' in fnu or '模' in fn or any(k in p for p in parts for k in ('模号',))
    if field == 'machine':
        return '#' in fn or re.search(r'\b(DC|DV|DT)\d', fn, re.I) is not None or \
               any(k in p for p in parts for k in ('机台号', '机台', '台号'))
    if field == 'test_type':
        return any(k in fn for k in ('首件', '制程', '尺寸')) or \
               any(k in p for p in parts for k in ('检测类型', '类型'))
    if field == 'send_date':
        return bool(re.search(r'\d+\.\d+', fn)) or any(k in p for p in parts for k in ('送测日期', '日期', '送测日'))
    if field == 'send_time':
        return bool(re.search(r'\d{1,2}[点:：]\d{0,2}', fn)) or \
               any(k in p for p in parts for k in ('送测时间', '时间', '送测时'))
    if field == 'quantity':
        return bool(re.search(r'PCS|个|件', fn, re.I)) or any(k in p for p in parts for k in ('数量', '数', '量'))
    return False


_MEASUREMENT_MARKER_RE = re.compile(r'^(?:O{1,3}M{1,3}|CMM|CM|C|O)(?:\d+#?)?$', re.IGNORECASE)
_MEASUREMENT_OPERATOR_RE = re.compile(
    r'(?:O{1,3}M{1,3}(?!M)|CMM)(?:\d+#?)?(?:[\s\-_]+)?([\u4e00-\u9fa5]{2,4}|[A-Za-z]{1,6})(?![#\d])',
    re.IGNORECASE,
)
_OPERATOR_STOP_WORDS = {
    'AOI', 'CNC', 'CMM', 'OMM', 'OM', 'CM', 'O', 'C',
    'M', 'PCS', 'FAI', 'IQC', 'OQC', 'BIN2', 'BIN4',
    '首件', '制程', '尺寸', '全尺寸', '射出', '烧结', '整形', '镭雕',
    '焊接', '开发', '测量', '量测', '送测', '检测', '测试', '手量',
    '对标', '对照', '校准', '标准',
}
_OMM_ORDINARY_OPERATORS = {
    '禹欣', '何淑畅', '赵亚琪', '卫阳', '王业陈', '金志豪',
    '于晚杰', '王卓越', '郑家午', '付成坤', '李晓冉', '魏则元',
}
_OMM_MANAGER_OPERATORS = {'王婷', '陈跃进'}
_OPERATOR_NAME_ALIASES = {
    '魏泽元': '魏则元',
}
_EXTRA_MEASUREMENT_PERSON_NAMES = {'蒋金潘', '张婉茹'}
_MEASUREMENT_PERSON_NAMES = (
    _OMM_ORDINARY_OPERATORS |
    _OMM_MANAGER_OPERATORS |
    _EXTRA_MEASUREMENT_PERSON_NAMES |
    set(_OPERATOR_NAME_ALIASES.keys())
)


def _is_measurement_marker_token(token):
    """OMM/CMM/O/C plus optional machine number such as OMM5#."""
    if not token:
        return False
    return bool(_MEASUREMENT_MARKER_RE.match(str(token).strip()))


def _measurement_marker_index(parts):
    for index, part in enumerate(parts):
        if _is_measurement_marker_token(part):
            return index
    return -1


def _measurement_people_aliases(measurement_people=None):
    aliases = dict(_OPERATOR_NAME_ALIASES)
    if isinstance(measurement_people, dict):
        for key, value in (measurement_people.get('aliases') or {}).items():
            key = str(key or '').strip()
            value = str(value or '').strip()
            if key and value:
                aliases[key] = value
    return aliases


def _measurement_person_names(measurement_people=None):
    names = set(_MEASUREMENT_PERSON_NAMES)
    if isinstance(measurement_people, dict):
        for key in ('ordinary', 'managers', 'extra', 'names'):
            for name in measurement_people.get(key) or []:
                clean = str(name or '').strip()
                if clean:
                    names.add(clean)
        for alias, canonical in (measurement_people.get('aliases') or {}).items():
            alias = str(alias or '').strip()
            canonical = str(canonical or '').strip()
            if alias:
                names.add(alias)
            if canonical:
                names.add(canonical)
    return names


def _has_measurement_people_config(measurement_people=None):
    if not isinstance(measurement_people, dict):
        return False
    if any(key in measurement_people for key in ('ordinary', 'managers', 'extra', 'names', 'aliases')):
        return True
    if measurement_people.get('aliases'):
        return True
    return any(measurement_people.get(key) for key in ('ordinary', 'managers', 'extra', 'names'))


def _is_valid_operator_name(name):
    if not name:
        return False
    name = str(name).strip()
    if not name or name.upper() in _OPERATOR_STOP_WORDS or name in _OPERATOR_STOP_WORDS:
        return False
    if re.search(r'\d|#|PCS', name, re.IGNORECASE):
        return False
    return bool(re.match(r'^[\u4e00-\u9fa5]{2,4}$', name) or re.match(r'^[A-Za-z]{1,6}$', name))


def _normalize_operator_name(name, measurement_people=None):
    name = str(name or '').strip()
    return _measurement_people_aliases(measurement_people).get(name, name)


def parse_folder_name(folder_name, parent_shift_suffix='', parent_date='', known_senders=None, measurement_people=None):
    result = {
        'station': '/', 'product': '/', 'sender': '/', 'work_order': '无工单号',
        'mold': '/', 'machine': '/', 'test_type': '/',
        'send_date': '/', 'send_time': '/', 'quantity_str': '/',
        'operator': '/',
    }

    parts = folder_name.split('-')

    def has_placeholder(*keywords):
        return any(kw in p for kw in keywords for p in parts)

    has_sender_placeholder = has_placeholder('送测人', '送测人员', '送测')
    has_work_order_placeholder = has_placeholder('工单号', '工单', '单号')
    has_mold_placeholder = has_placeholder('模号', '模')
    has_machine_placeholder = has_placeholder('机台号', '机台', '台号')
    has_date_placeholder = has_placeholder('送测日期', '日期', '送测日')
    has_time_placeholder = has_placeholder('送测时间', '时间', '送测时')
    has_quantity_placeholder = has_placeholder('数量', '数', '量')
    has_test_type_placeholder = has_placeholder('检测类型', '类型')

    folder_upper = folder_name.upper()
    if 'FAI' in folder_upper or any(k in folder_name for k in ('测试片', '测试', '孔内直径', '外径', '内径', '直径', '圆度', 'cam件')):
        result['station'] = '开发'
    elif 'CNC' in folder_name:
        result['station'] = 'CNC'
    elif '射出' in folder_name:
        result['station'] = '射出'
    elif '镭雕' in folder_name or 'BIN2' in folder_name or 'BIN4' in folder_name:
        result['station'] = '镭雕'
    elif '整形' in folder_name:
        result['station'] = '整形'
    elif '烧结' in folder_name:
        result['station'] = '烧结'
    elif '焊接' in folder_name:
        result['station'] = '焊接'
    elif 'IQC' in folder_upper:
        result['station'] = 'IQC'
    elif 'OQC' in folder_upper:
        result['station'] = 'OQC'
    elif re.search(r'\b[A-Z]QC\b', folder_upper):
        result['station'] = re.search(r'\b([A-Z]QC)\b', folder_upper).group(1)

    # 没有任何生产工站特征时，默认归为开发，避免大量无意义弹窗
    if result['station'] == '/' and not _field_strong_pattern_detected('station', folder_name, parts):
        result['station'] = '开发'

    # 开发/IQC/OQC/XQC 等特殊工站通常没有机台号，不强制缺失
    is_dev = result['station'] in ('开发', 'IQC', 'OQC') or re.match(r'^[A-Z]QC$', result['station'])

    # 产品：焊接/AOI 口径下，613/41424~41429 不是日报品名，应优先取括号内 034~039。
    is_welding_aoi = result['station'] == '焊接' or '焊接' in folder_name or 'AOI' in folder_name.upper()
    if is_welding_aoi:
        m = re.search(r'\((03[4-9])-\d{3}\)', folder_name)
        if m:
            result['product'] = m.group(1)
        else:
            m = re.search(r'(?<!\d)(289|290)(?!\d)', folder_name)
            if m:
                result['product'] = m.group(1)

    # 普通产品：优先识别开头三位数字（如 035-CNC...），否则从五位数取后三位
    if result['product'] == '/':
        m = re.match(r'^(\d{3})\b', folder_name)
        if m:
            result['product'] = m.group(1)
        else:
            m = re.search(r'(\d{5})(?:-\d{2})', folder_name)
            if m:
                result['product'] = m.group(1)[-3:]
            else:
                m2 = re.search(r'(\d{5})', folder_name)
                if m2:
                    result['product'] = m2.group(1)[-3:]

    # 测量人员：OMM/CMM 后面必须是明确人名/首字母；OMM4#/CMM5# 仅为机台号。
    # 支持 OMM4#-卫阳、OMM5#卫阳、OMM-卫阳、CMM5#-卫阳 等写法。
    operator_matches = []
    strict_measurement_people = _has_measurement_people_config(measurement_people)
    measurement_person_names = _measurement_person_names(measurement_people)
    for op_match in _MEASUREMENT_OPERATOR_RE.finditer(folder_name):
        operator_candidate = _normalize_operator_name(op_match.group(1), measurement_people)
        if strict_measurement_people and operator_candidate not in measurement_person_names:
            continue
        if _is_valid_operator_name(operator_candidate):
            operator_matches.append(operator_candidate)
    if operator_matches:
        result['operator'] = operator_matches[-1]

    # 送测人识别（多策略）
    # 常见工站名/检测类型/工序名，不能当成人名
    _STATION_WORDS = {
        '首件', '制程', '尺寸', '全尺寸', '射出', '烧结', '整形', '镭雕',
        '焊接', 'AOI', 'CNC', 'CMM', 'OMM', '开发', '测量', '量测',
        '送测', '生成', '制作', '数量', '复测', '检测', '测试',
        '送测人员', '送测人', '量测人员', 'CMM量测人员', 'OMM量测人员',
        '手量', '手动量测', '手工量测', 'O', 'C', 'OM', 'CM',
        '人员', '对标', '对照', '校准', '标准',
    }
    _NON_SENDER_TOKENS = (
        '对标', '对照', '校准', '标准', '焊接', '量测', '人员',
        '尺寸', '全尺寸', '测试片', '检测', '测试', '手量',
    )
    _DEFAULT_KNOWN_SENDER_HINTS = ('张元庆',)
    # 送测关键词（"送测"及常见误打）
    _SENDER_KEYWORDS = r'(?:送测|生成|制作|送检|提交|上交|填写)'

    normalized_for_sender = folder_name.replace('：', ':').replace('；', ';')
    # 先剥离"送测人员"/"送测人"前缀，避免"员彭立娜"被误匹配
    normalized_for_sender = re.sub(r'送测人员?', '', normalized_for_sender, count=1)
    known_sender_list = []

    def _is_valid_sender(name):
        """判断识别到的候选是否是有效人名（非工站名等）。"""
        if not name:
            return False
        name = name.strip()
        if not name or name in _STATION_WORDS or name.upper() in _STATION_WORDS:
            return False
        if any(token in name for token in _NON_SENDER_TOKENS):
            return False
        # 纯中文 2-4 字
        if re.match(r'^[\u4e00-\u9fa5]{2,4}$', name):
            return True
        # 英文名 2-20 字母（可含点和连字符）
        if re.match(r'^[A-Za-z][A-Za-z.\-]{1,19}$', name):
            return True
        return False

    seen_known = set()
    for sender in list(known_senders or []) + list(_DEFAULT_KNOWN_SENDER_HINTS):
        sender = str(sender).strip()
        if sender and sender not in seen_known and _is_valid_sender(sender):
            seen_known.add(sender)
            known_sender_list.append(sender)
    known_sender_list.sort(key=len, reverse=True)

    def _match_known_sender_near_action():
        """只在动作词附近使用词库，避免把 CMM/OMM 后的测量员当成送测人。"""
        bridge = r'[\s\-_.:：;；\d点分ABab年月日]*'
        for sender in known_sender_list:
            escaped = re.escape(sender)
            if re.search(escaped + bridge + _SENDER_KEYWORDS, normalized_for_sender, re.IGNORECASE):
                return sender
            if re.search(_SENDER_KEYWORDS + r'(?:人员|人)?' + bridge + escaped, normalized_for_sender, re.IGNORECASE):
                return sender
        return None

    def _match_known_sender_without_action():
        """无"送测"动作词时，已知送测人可作为候选，但避开明确量测员上下文。"""
        role_tokens = (
            '量测人员', '测量员', '测量人员', 'CMM量测人员', 'OMM量测人员',
            '手量人员', '手动量测', '手工量测',
        )
        tool_tokens = ('OMM', 'OM', 'O', 'CMM', 'CM', 'C', '手量')
        for sender in known_sender_list:
            if sender == result.get('operator'):
                continue
            if sender in measurement_person_names:
                continue
            for idx, part in enumerate(parts):
                candidate = part.strip()
                if candidate != sender:
                    continue
                if candidate in measurement_person_names:
                    continue
                prev_part = parts[idx - 1].strip() if idx > 0 else ''
                next_part = parts[idx + 1].strip() if idx + 1 < len(parts) else ''
                if any(token in candidate or token in prev_part or token in next_part for token in role_tokens):
                    continue
                if _is_measurement_marker_token(prev_part) or _is_measurement_marker_token(next_part):
                    continue
                if candidate.upper() in tool_tokens or _is_measurement_marker_token(candidate):
                    continue
                if _is_valid_sender(candidate):
                    return candidate
        return None

    known_sender_match = _match_known_sender_near_action()
    if known_sender_match:
        result['sender'] = known_sender_match

    # 策略1：姓名(时间)?送测关键词 —— "张三送测" / "张三14:00送测" / "张三生成"
    sender_match = re.search(
        r'([\u4e00-\u9fa5]{2,4}|[A-Za-z]{2,20})(?:\d{1,2}[:.点]\d{0,2}(?:分)?)?' + _SENDER_KEYWORDS,
        normalized_for_sender
    )
    if result['sender'] == '/' and sender_match and _is_valid_sender(sender_match.group(1)):
        result['sender'] = sender_match.group(1)
    elif result['sender'] == '/':
        # 策略2：送测关键词?姓名(时间)? —— "送测人员彭立娜送测" / "送测张三"
        sender_match2 = re.search(
            _SENDER_KEYWORDS + r'(?:人员|人)?([\u4e00-\u9fa5]{2,4}|[A-Za-z]{2,20})',
            normalized_for_sender
        )
        if sender_match2 and _is_valid_sender(sender_match2.group(1)):
            result['sender'] = sender_match2.group(1)

    if result['sender'] == '/':
        # 策略3：时间+姓名（无关键词） —— "22:40张三" / "4:00刘前程送测" / "6.5B周琳4:20送测"
        # 匹配 时间(可选分隔符)中文姓名 的模式
        m_time_name = re.search(
            r'\d{1,2}[:.]\d{0,2}(?:分)?[\s\-]*([\u4e00-\u9fa5]{2,4})',
            normalized_for_sender
        )
        if m_time_name and _is_valid_sender(m_time_name.group(1)):
            result['sender'] = m_time_name.group(1)

    if result['sender'] == '/':
        # 策略4：姓名-PCS数量 —— "陈大勇-16PCS" / "陈大勇16PCS" / "张三-数量"
        m_name_pcs = re.search(
            r'([\u4e00-\u9fa5]{2,4})(?:[\-]?\d+PCS|[\-]?\d+pcs|[\-]?数量)',
            normalized_for_sender, re.IGNORECASE
        )
        if m_name_pcs and _is_valid_sender(m_name_pcs.group(1)):
            result['sender'] = m_name_pcs.group(1)

    if result['sender'] == '/' and known_sender_list:
        known_sender_match = _match_known_sender_without_action()
        if known_sender_match:
            result['sender'] = known_sender_match

    if result['sender'] == '/':
        # 策略5：回退 —— 在 OMM/CMM 之前找中文或英文名，跳过工站名和短代码
        omm_idx = _measurement_marker_index(parts)
        cmm_idx = next((i for i, p in enumerate(parts) if str(p).strip().upper().startswith('CMM')), -1)
        search_end = omm_idx
        if cmm_idx != -1 and (omm_idx == -1 or cmm_idx < omm_idx):
            search_end = cmm_idx
        if search_end > 0:
            for j in range(search_end - 1, -1, -1):
                candidate = parts[j].strip()
                if candidate in ('数量',) or candidate == result['operator']:
                    continue
                if _is_valid_sender(candidate):
                    # 额外检查：候选里不能含数字/PCS/#等
                    if not re.search(r'\d|PCS|#|ww|WW', candidate, re.IGNORECASE):
                        result['sender'] = candidate
                        break

    # 工单号
    m = re.search(r'(ww\d+-\d+)', folder_name, re.IGNORECASE)
    if m:
        result['work_order'] = m.group(1).upper()
    elif '无工单号' in folder_name:
        result['work_order'] = '无工单号'

    # 模号：优先识别显式模号；CNC 工站未识别到时默认 T1
    m_bin = re.search(r'(BIN[24])', folder_name, re.IGNORECASE)
    m_code = re.search(r'-(M[a-zA-Z0-9]+)-', folder_name)
    if not m_code:
        m_code = re.search(r'-([DT]\d+)-', folder_name)
    if not m_code:
        m_code = re.search(r'-(\d+模)-', folder_name)
    if m_bin:
        result['mold'] = m_bin.group(1).upper()
    elif m_code:
        result['mold'] = m_code.group(1).upper() if '模' not in m_code.group(1) else m_code.group(1)
    elif result['station'] == 'CNC':
        result['mold'] = 'T1'

    # 机台号
    m = re.search(r'([A-Z0-9]+#)', folder_name)
    if not m:
        m = re.search(r'-((?:DC|DV)\d{2})-', folder_name)
    if not m:
        # D0E3、T1、T2 等开发/特殊机台号
        m = re.search(r'-([DT][A-Z0-9]{1,4})-', folder_name)
    if m:
        result['machine'] = m.group(1)

    # 检测类型
    if '首件' in folder_name:
        result['test_type'] = '首件'
    elif '制程' in folder_name:
        result['test_type'] = '制程'
    elif '尺寸' in folder_name:
        result['test_type'] = '尺寸'

    # 日期
    m = re.search(r'(\d+\.\d+[a-zA-Z]?)', folder_name)
    if m:
        result['send_date'] = m.group(1)

    # 如果文件夹名里没有日期，尝试使用父文件夹日期
    if result['send_date'] == '/' and parent_date:
        result['send_date'] = parent_date

    has_shift = result['send_date'][-1].upper() in 'AB' if result['send_date'] != '/' else False
    if parent_shift_suffix and result['send_date'] != '/' and not has_shift:
        result['send_date'] += parent_shift_suffix

    # 送测时间
    m = re.search(r'\d+\.\d+[a-zA-Z]?-([\d.：:；;点]+?)-', folder_name)
    if not m:
        m = re.search(r'\d+\.\d+[a-zA-Z]?-(\d{1,2}[点:]\d{0,2}(?:分)?)(?=[\u4e00-\u9fa5A-Za-z])', folder_name.replace('：', ':'))
    if not m:
        m = re.search(r'-(\d{1,2}点\d{0,2}(?:分)?)-', folder_name)
    if not m:
        # 刘前程2点30送测 / 彭立娜14:00送测 这类无前后横杠的时间
        m = re.search(r'(\d{1,2}[点:]\d{0,2}(?:分)?)送测', folder_name.replace('：', ':'))
    if not m:
        # 7:55张颖龙送测 / 7点55张颖龙送测 这类时间紧贴送测人的写法
        m = re.search(r'(\d{1,2}[点:]\d{0,2}(?:分)?)[\s\-]*[\u4e00-\u9fa5]{2,4}送测', folder_name.replace('：', ':'))
    if m:
        candidate = m.group(1).replace('：', ':').replace('；', ';').replace('点', ':')
        if re.match(r'^\d{1,2}[:;]\d{2}$', candidate) or re.match(r'^\d+\.\d+$', candidate):
            result['send_time'] = candidate
        elif re.match(r'^\d{1,2}:$', candidate):
            result['send_time'] = candidate + '00'
        elif re.match(r'^\d{1,2}:\d{1,2}分$', candidate):
            result['send_time'] = candidate.rstrip('分')

    # 数量
    m = re.search(r'(\d+)\s*PCS', folder_name, re.IGNORECASE)
    if m:
        result['quantity_str'] = m.group(1)
    else:
        # 回退：看 OMM/CMM 左侧紧邻的纯 1~2 位数字（如 ...-12-omm-...）
        normalized = re.sub(r'(?i)(omm|cmm)([\u4e00-\u9fa5]{2,4})', r'\1-\2', folder_name)
        qty_parts = normalized.split('-')
        omm_idx = next((i for i, p in enumerate(qty_parts) if p.upper() == 'OMM'), -1)
        cmm_idx = next((i for i, p in enumerate(qty_parts) if p.upper() == 'CMM'), -1)
        target_idx = -1
        if omm_idx != -1 and cmm_idx != -1:
            target_idx = min(omm_idx, cmm_idx)
        elif omm_idx != -1:
            target_idx = omm_idx
        elif cmm_idx != -1:
            target_idx = cmm_idx

        qty_from_pos = None
        if target_idx > 0:
            left = qty_parts[target_idx - 1]
            if re.match(r'^\d{1,2}$', left):
                qty_from_pos = left

        if qty_from_pos:
            result['quantity_str'] = qty_from_pos

    # ==================== 置信度模型：区分"确实没有"和"识别失败" ====================
    missing = []
    placeholders = []

    # 关键字段：未识别即需要人工确认
    if result['sender'] == '/':
        if has_sender_placeholder:
            placeholders.append('sender')
        else:
            missing.append('sender')

    if result['quantity_str'] == '/':
        if has_quantity_placeholder:
            placeholders.append('quantity')
        elif result['station'] == 'CNC':
            # CNC 固定 20 分钟，没写数量时直接标 "/"，不弹窗
            result['quantity_str'] = '/'
        else:
            missing.append('quantity')

    # 非关键字段：只有强模式存在却识别失败时才弹窗；完全没有相关模式视为"确实没有"
    has_explicit_no_wo = any('无工单号' in p for p in parts)
    if result['work_order'] == '无工单号':
        if has_work_order_placeholder and not has_explicit_no_wo:
            placeholders.append('work_order')
        elif _field_strong_pattern_detected('work_order', folder_name, parts) and not has_explicit_no_wo:
            placeholders.append('work_order')

    if result['mold'] == '/':
        if has_mold_placeholder:
            placeholders.append('mold')
        elif _field_strong_pattern_detected('mold', folder_name, parts):
            placeholders.append('mold')

    if result['machine'] == '/':
        if has_machine_placeholder:
            placeholders.append('machine')
        elif _field_strong_pattern_detected('machine', folder_name, parts):
            placeholders.append('machine')

    if result['test_type'] == '/':
        if has_test_type_placeholder:
            placeholders.append('test_type')
        elif _field_strong_pattern_detected('test_type', folder_name, parts):
            placeholders.append('test_type')

    if result['send_date'] == '/':
        if has_date_placeholder:
            placeholders.append('send_date')
        elif _field_strong_pattern_detected('send_date', folder_name, parts):
            missing.append('send_date')

    if result['send_time'] == '/':
        if has_time_placeholder:
            placeholders.append('send_time')
        elif _field_strong_pattern_detected('send_time', folder_name, parts):
            placeholders.append('send_time')

    if result['station'] == '/':
        if _field_strong_pattern_detected('station', folder_name, parts):
            missing.append('station')

    if result['product'] == '/':
        if _field_strong_pattern_detected('product', folder_name, parts):
            missing.append('product')

    return result, missing, placeholders


def _is_standard_template(ws):
    """判断 worksheet 是否为公司标准首件尺寸报告模板。

    标准特征：
    - 第一行（或前 3 行）包含"安徽中耀智能科技有限公司"和"首件尺寸报告"
    - 存在"测量值"单元格
    """
    header_text = ''
    for row in range(1, min(4, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(row=row, column=col).value
            if v:
                header_text += str(v)

    has_company = '安徽中耀智能科技有限公司' in header_text
    has_report_title = '首件尺寸报告' in header_text

    has_measure_value = False
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(row=row, column=col).value
            if v and '测量值' in str(v):
                has_measure_value = True
                break
        if has_measure_value:
            break

    return has_company and has_report_title and has_measure_value


def _count_sheet_quantity(ws):
    """统计标准模板中实际测量的件数（按列统计，每列存在有效 OMM 测量数据即算 1 件）。

    模板理解：
    - 每个 sheet 中可能有一个主测量表，也可能纵向复制多段同结构测量表。
    - EVT/全尺寸等模板优先以短"检测工具"列锚定，只统计检测工具为 OMM 的行。
    - 测量列头可能是 1-1、1-2、2-1、2-2，不能用前缀数字递增/回退判断边界。
    - 部分模板（如 CNC）没有"检测工具"列，直接统计主测量表中有数据的列。
    - 列头含复测/掉料/丢料/丢失/损坏/残废/报废/缺料等字样时跳过。

    核心逻辑：
    1. 若存在短"检测工具"列，按每个"检测工具"锚点拆成独立测量块，统计 OMM 行下有数值的列并累加。
    2. 否则定位"测量值"单元格，按旧模板方式估算主测量表区域。
    3. 每列有数据即 1 件。

    如果传入的不是标准模板，返回 0，调用方应提示"非标准表格无法识别"。
    """

    if not _is_standard_template(ws):
        return 0

    INVALID_HEADER_MARKS = ('复测', '掉料', '丢料', '丢失', '损坏', '残废', '报废', '缺料')

    def _is_numeric_data(val):
        if val is None:
            return False
        if isinstance(val, (int, float)):
            return True
        s = str(val).strip()
        if not s or s in ('送测', '复测', '测量值', '合格', '不合格', 'OK', 'NG'):
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False

    def _header_invalid(val):
        if val is None:
            return False
        s = str(val).strip()
        return any(k in s for k in INVALID_HEADER_MARKS)

    def _is_omm_tool(val):
        if val is None:
            return False
        s = str(val).strip().upper()
        return 'OMM' in s and 'CMM' not in s

    def _is_tool_marker(val):
        if val is None:
            return False
        s = str(val).strip().upper()
        return s in ('OMM', 'CMM') or s.startswith('OMM') or s.startswith('CMM')

    def _compact_text(val):
        if val is None:
            return ''
        return re.sub(r'\s+', '', str(val).strip())

    def _valid_measure_header(val):
        s = _compact_text(val)
        if not s:
            return False
        if any(k in s for k in INVALID_HEADER_MARKS):
            return False
        if any(k in s for k in ('检测工具', '检测结果', '测量值', '判断', '量测分析', '合格', '不合格')):
            return False
        # 真正的测量列头通常是 1、1-1、2-2、01 等纯数字/短横组合；中文列头不参与件数统计。
        return bool(re.match(r'^\d+(?:-\d+)?$', s))

    def _is_measurement_block_boundary(val):
        s = _compact_text(val)
        if not s:
            return False
        # "检测结果/测量值"是测量区本身；"量测分析/判断"开始后面的分析区，不参与件数统计。
        return any(k in s for k in ('量测分析', '判断', '判定', '备注'))

    # ── 优先路径：短"检测工具"列锚定 ──
    # 同一个 sheet 里可能纵向复制多个测量表（如首件 FAI 表格上半段 7 件、下半段 4 件）。
    # 不能只选一个最佳"检测工具"锚点；每段应独立统计后累加。
    tool_headers = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 80)):
            text = _compact_text(ws.cell(row=row, column=col).value)
            if '检测' not in text or '工具' not in text:
                continue
            tool_headers.append((row, col))

    if tool_headers:
        block_counts = []
        for tool_header_row, tool_col in sorted(set(tool_headers)):
            next_header_rows = [
                row for row, col in tool_headers
                if col == tool_col and row > tool_header_row
            ]
            block_end_row = (min(next_header_rows) - 1) if next_header_rows else min(ws.max_row, tool_header_row + 500)

            tool_rows = [
                row for row in range(tool_header_row + 1, block_end_row + 1)
                if _is_tool_marker(ws.cell(row=row, column=tool_col).value)
            ]
            if not tool_rows:
                continue

            first_tool_row = min(tool_rows)
            header_candidates = []
            for row in range(tool_header_row + 1, first_tool_row):
                measure_end_col = ws.max_column
                for col in range(tool_col + 1, ws.max_column + 1):
                    if any(
                        _is_measurement_block_boundary(ws.cell(row=header_row, column=col).value)
                        for header_row in range(tool_header_row, row)
                    ):
                        measure_end_col = col - 1
                        break
                valid_cols = [
                    col for col in range(tool_col + 1, measure_end_col + 1)
                    if _valid_measure_header(ws.cell(row=row, column=col).value)
                ]
                if valid_cols:
                    header_candidates.append((len(valid_cols), row, valid_cols))

            if not header_candidates:
                continue

            _, id_row, measure_cols = max(header_candidates, key=lambda item: item[0])
            omm_rows = [
                row for row in tool_rows
                if _is_omm_tool(ws.cell(row=row, column=tool_col).value)
            ]
            if not omm_rows:
                continue

            count = 0
            for col in measure_cols:
                header_val = ws.cell(row=id_row, column=col).value
                if _header_invalid(header_val):
                    continue
                if any(_is_numeric_data(ws.cell(row=row, column=col).value) for row in omm_rows):
                    count += 1
            if count > 0:
                block_counts.append(count)

        if block_counts:
            return sum(block_counts)

    # ── 步骤 1：定位主测量表区域 ──
    measure_row = measure_col = None
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(30, ws.max_column + 1)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and '测量值' in str(cell_val):
                measure_row, measure_col = row, col
                break
        if measure_row:
            break

    if not measure_row or not measure_col:
        return None

    # 自动定位编号行（可能在测量值下方若干行内）
    id_row = None
    for row in range(measure_row + 1, min(measure_row + 6, ws.max_row + 1)):
        v = ws.cell(row=row, column=measure_col).value
        if v is not None and str(v).strip() != '':
            id_row = row
            break

    if id_row is None:
        return None

    data_start_row = id_row + 1
    if data_start_row > ws.max_row:
        return None

    def _id_to_int(val):
        if val is None:
            return None
        s = str(val).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except (ValueError, TypeError):
            pass
        m = re.match(r'^(\d+)', s)
        if m:
            try:
                return int(m.group(1))
            except (ValueError, TypeError):
                pass
        return None

    # 确定主测量表的右边界：
    # - 标题行（measure_row）出现新的非"测量值"表头
    # - 编号行出现重复/回退
    # - 编号行连续空列
    last_id = -1
    end_col = ws.max_column
    for col in range(measure_col, ws.max_column + 1):
        header_val = ws.cell(row=measure_row, column=col).value
        if col > measure_col and header_val is not None:
            hs = str(header_val).strip()
            if hs and hs != '测量值':
                end_col = col - 1
                break

        id_val = ws.cell(row=id_row, column=col).value
        if id_val is None or str(id_val).strip() == '':
            end_col = col - 1
            break

        nid = _id_to_int(id_val)
        if nid is not None and nid <= last_id:
            end_col = col - 1
            break
        if nid is not None:
            last_id = nid

    # ── 步骤 2：查找检测工具列（可选）──
    tool_col = None
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 60)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                s = str(cell_val)
                if '检测' in s and '工具' in s:
                    tool_col = col
                    break
        if tool_col:
            break

    # 收集 OMM 行（如果存在检测工具列）
    omm_rows = set()
    if tool_col:
        for row in range(data_start_row, min(ws.max_row + 1, data_start_row + 500)):
            if _is_omm_tool(ws.cell(row=row, column=tool_col).value):
                omm_rows.add(row)

    # ── 步骤 3：在主测量表区域内逐列统计 ──
    count = 0
    for col in range(measure_col, end_col + 1):
        id_val = ws.cell(row=id_row, column=col).value
        if id_val is None or str(id_val).strip() == '':
            break
        if _header_invalid(id_val):
            continue

        has_data = False
        if omm_rows:
            for row in omm_rows:
                if _is_numeric_data(ws.cell(row=row, column=col).value):
                    has_data = True
                    break
        else:
            for row in range(data_start_row, min(ws.max_row + 1, data_start_row + 200)):
                if _is_numeric_data(ws.cell(row=row, column=col).value):
                    has_data = True
                    break

        if has_data:
            count += 1

    return count if count > 0 else None


def read_xlsx_quantity(xlsx_path):
    """读取 xlsx 中实际测量的件数；若非标准表格，返回 0。

    返回值：
    - int > 0：实际测量件数
    - 0：存在 sheet，但非标准模板或没有可识别的有效数据
    - None：读取文件失败
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"读取xlsx出错: {e}")
        return None

    total = 0
    has_standard_sheet = False
    for ws in wb.worksheets:
        try:
            if _is_standard_template(ws):
                has_standard_sheet = True
                qty = _count_sheet_quantity(ws)
                if qty:
                    total += qty
        except Exception as e:
            print(f"统计sheet出错 ({ws.title}): {e}")
            continue
    wb.close()

    if not has_standard_sheet:
        return 0
    return total if total > 0 else 0


def _parse_time_str(s):
    """将 '9:30'/'09:30'/'9.30' 等字符串转为 datetime.time, 失败返回 None"""
    if not s or s == '/':
        return None
    s = str(s).strip().replace('：', ':').replace('；', ';').replace('.', ':')
    m = re.match(r'^(\d{1,2}):(\d{2})$', s)
    if m:
        h, minute = int(m.group(1)), int(m.group(2))
        if 0 <= h < 24 and 0 <= minute < 60:
            return dtime(h, minute)
    return None


def _get_operator_initials(name):
    """获取姓名的拼音首字母；仅对中文有效，非中文返回小写原串。"""
    if not name:
        return ''
    # 纯字母名字直接小写返回
    if re.match(r'^[A-Za-z]+$', name):
        return name.lower()
    try:
        from pypinyin import lazy_pinyin
        return ''.join(p[0].lower() for p in lazy_pinyin(name) if p)
    except Exception:
        # 无 pypinyin 时，用内置简单映射兜底
        _SIMPLE_INITIALS = {
            '禹欣': 'yx',
        }
        return _SIMPLE_INITIALS.get(name, name.lower())


def _operator_name_candidates(name, measurement_people=None):
    candidates = {name}
    aliases = _measurement_people_aliases(measurement_people)
    for alias, canonical in aliases.items():
        if canonical == name:
            candidates.add(alias)
    return candidates


def _make_operator_patterns(name, measurement_people=None):
    """根据中文姓名生成多种容错匹配正则：中文、首字母、OMM/OM、-可有可无、大小写不敏感。"""
    patterns = []
    candidates = _operator_name_candidates(name, measurement_people)
    initials = _get_operator_initials(name)
    if initials:
        candidates.add(initials)

    for cand in candidates:
        # O+MM? 表示 O 后面跟 1~3 个 M 都算（容错 OM/OMM/OOMM/OMMM）
        # OMM4#/CMM5# 中间的数字和 # 是机台号，不影响后面测量员匹配
        patterns.append(rf'(?:O{{1,3}}M{{1,3}}|CMM)(?:\d+#?)?[\-_]?{re.escape(cand)}')
    return re.compile('|'.join(patterns), re.IGNORECASE)


def _is_operator_folder(folder_name, operator_name, measurement_people=None):
    """判断文件夹名是否匹配指定操作者（支持中文/首字母/OM/OMM/连接符容错，大小写不敏感）。"""
    if not operator_name:
        operator_name = '禹欣'
    pattern = _make_operator_patterns(operator_name, measurement_people)
    return bool(pattern.search(folder_name))


def _is_untitled_folder(folder_name):
    return bool(re.match(r'^\s*新建文件夹(?:\s*\(\d+\))?\s*$', str(folder_name or '')))


def _has_reviewable_payload(folder_path):
    try:
        for name in os.listdir(folder_path):
            if os.path.isfile(os.path.join(folder_path, name)):
                return True
    except OSError:
        return False
    return False


def _contains_measurement_marker(folder_name):
    parts = re.split(r'[\s\-_]+', str(folder_name or ''))
    return any(_is_measurement_marker_token(part) for part in parts)


def parse_all_folders(base_dir, operator_name=OPERATOR_NAME, known_senders=None, measurement_people=None):
    """解析 base_dir 下的所有子文件夹（忽略文件），按 OMM-操作者 过滤。"""
    items = os.listdir(base_dir)
    folders = sorted(f for f in items if os.path.isdir(os.path.join(base_dir, f)))
    skipped_files = [f for f in items if not os.path.isdir(os.path.join(base_dir, f))]
    if skipped_files:
        print(f"[跳过非文件夹项] {len(skipped_files)} 个: {skipped_files[:5]}{'...' if len(skipped_files) > 5 else ''}")
    parent_name = os.path.basename(base_dir.rstrip('\\/'))
    shift_suffix = parent_name[-1] if parent_name and parent_name[-1].upper() in 'AB' else ''
    # 从父文件夹名提取日期（如 6.11B → 6.11B，6.11 → 6.11），用于子文件夹缺失日期时回退
    date_match = re.match(r'(\d+)\.(\d+)[AB]?', parent_name)
    parent_date = date_match.group(0) if date_match else ''
    records = []
    review_map = {}  # folder -> {'missing': [...], 'placeholders': [...]}
    for folder in folders:
        folder_path = os.path.join(base_dir, folder)
        xlsx_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
        needs_untitled_review = _is_untitled_folder(folder) and (xlsx_files or _has_reviewable_payload(folder_path))
        parsed, missing, placeholders = parse_folder_name(
            folder,
            parent_shift_suffix=shift_suffix,
            parent_date=parent_date,
            known_senders=known_senders,
            measurement_people=measurement_people,
        )
        has_payload = bool(xlsx_files) or _has_reviewable_payload(folder_path)
        needs_operator_review = parsed.get('operator') == '/' and _contains_measurement_marker(folder) and has_payload
        if not _is_operator_folder(folder, operator_name, measurement_people) and not needs_untitled_review and not needs_operator_review:
            print(f"跳过（非{operator_name}）: {folder}")
            continue
        review_warnings = []
        if needs_untitled_review:
            review_warnings.append('疑似未命名资料目录：目录内存在文件，请补全识别信息，或在审核弹窗中选择忽略该目录。')
            for field in ('station', 'product', 'sender', 'operator'):
                if field not in missing and field not in placeholders:
                    missing.append(field)
        if needs_operator_review:
            review_warnings.append('识别到 OMM/CMM 机台号但没有明确测量员；请补全测量员，或在审核弹窗中忽略该目录。')
            if 'operator' not in missing and 'operator' not in placeholders:
                missing.append('operator')
        review_map[folder] = {'missing': missing, 'placeholders': placeholders, 'warnings': review_warnings}

        # 数量优先级：文件夹名 PCS > xlsx 实际计数 > 缺失/弹窗
        actual_quantity = None
        nonstandard_xlsx = False
        qs = parsed['quantity_str']
        if qs != '/' and qs.isdigit():
            quantity = int(qs)  # 文件夹名优先
            if xlsx_files:
                actual_quantity = read_xlsx_quantity(os.path.join(folder_path, xlsx_files[0]))
                if actual_quantity is not None and actual_quantity > 0 and actual_quantity != quantity:
                    review_warnings.append(
                        f'Excel 读取件数({actual_quantity})与文件夹件数({quantity})不一致，默认采用文件夹件数。'
                    )
        else:
            if xlsx_files:
                actual_quantity = read_xlsx_quantity(os.path.join(folder_path, xlsx_files[0]))
            if actual_quantity is not None and actual_quantity > 0:
                quantity = actual_quantity
                # xlsx 已统计出数量，不再因数量缺失弹窗
                if 'quantity' in missing:
                    missing.remove('quantity')
                if 'quantity' in placeholders:
                    placeholders.remove('quantity')
            elif actual_quantity == 0:
                # xlsx 存在但非标准模板：标记为占位符，弹窗提示
                quantity = '/'
                nonstandard_xlsx = True
                if 'quantity' not in placeholders and 'quantity' not in missing:
                    placeholders.append('quantity')
            else:
                quantity = '/'
        records.append({
            'folder': folder, 'station': parsed['station'], 'product': parsed['product'],
            'sender': parsed['sender'], 'work_order': parsed['work_order'],
            'mold': parsed['mold'], 'machine': parsed['machine'],
            'test_type': parsed['test_type'], 'send_date': parsed['send_date'],
            'send_time': parsed['send_time'], 'quantity': quantity,
            'actual_quantity': actual_quantity,
            'nonstandard_xlsx': nonstandard_xlsx,
            'operator': parsed['operator'],
            'manual_duration': None,
            'manual_quantity': None,
            'missing_fields': missing,
            'placeholder_fields': placeholders,
            'reviewed': False,
        })
    return records, review_map


def schedule_tasks(records, shift_label='A', early_leave=False,
                   leave_strategy=None,
                   tpp_min=TPP_MIN_DEFAULT, tpp_max=TPP_MAX_DEFAULT,
                   pkg_rest=PKG_REST_DEFAULT,
                   enable_hand=True, enable_other=False,
                   special_items=None, duration_rules=None,
                   hand_max=120, other_max=90,
                   real_manual_tasks=None, filler_position='middle',
                   other_note='其他事务'):
    """排程任务。

    参数：
        leave_strategy: 'auto' | 'early' | 'normal' | None
            兼容旧参数 early_leave：
            - leave_strategy='early' 等价于 early_leave=True
            - leave_strategy='normal' 等价于 early_leave=False
            - leave_strategy='auto' 为智能判断（不自动下早班）
            - None 时退回到 early_leave 的 boolean 语义
        real_manual_tasks: 真实手量任务列表，作为独立任务参与排程并写入 Excel。
    """
    # 兼容旧参数
    if leave_strategy is None:
        leave_strategy = 'early' if early_leave else 'normal'
    if leave_strategy not in ('auto', 'early', 'normal'):
        leave_strategy = 'normal'
    if filler_position not in ('head', 'middle', 'tail', 'random'):
        filler_position = 'middle'
    other_note = str(other_note or '').strip() or '其他事务'

    cfg = SHIFTS[shift_label]
    offset = cfg['start_offset']
    schedule_warnings = []

    # tpp 下限 1.5 分钟
    original_tpp_min = tpp_min
    original_tpp_max = tpp_max
    if tpp_min < MIN_TPP:
        tpp_min = MIN_TPP
    if tpp_max < MIN_TPP:
        tpp_max = MIN_TPP
    if original_tpp_min < MIN_TPP or original_tpp_max < MIN_TPP:
        schedule_warnings.append(
            f'每件最低耗时不能低于{MIN_TPP}分钟，已自动按{MIN_TPP}分钟计算')
    if tpp_max < tpp_min:
        schedule_warnings.append(
            f'每件耗时上限({original_tpp_max}分钟)低于下限({original_tpp_min}分钟)，已按{tpp_min}分钟计算')
        tpp_max = tpp_min
    avg_tpp = (tpp_max + tpp_min) / 2

    # 规范化旧版固定件时物品列表，仅用于兼容迁移前配置
    _special_items = []
    if special_items:
        for it in special_items:
            if isinstance(it, dict):
                name = (it.get('name') or '').strip()
                try:
                    minutes = float(it.get('minutes', 8))
                except (TypeError, ValueError):
                    minutes = 8.0
                if name and minutes > 0:
                    _special_items.append({'name': name, 'minutes': minutes})

    def _match_special(rec):
        haystacks = []
        for k in ('folder', 'folder_name', 'station', 'product', 'test_type'):
            v = rec.get(k)
            if isinstance(v, str) and v and v != '/':
                haystacks.append(v)
        for it in _special_items:
            for h in haystacks:
                if it['name'] in h:
                    return it['minutes']
        return None

    _duration_rules = []
    if duration_rules:
        for rule in duration_rules:
            if not isinstance(rule, dict):
                continue
            if rule.get('enabled') is False or rule.get('deprecated') is True:
                continue
            matchers = rule.get('matchers') or []
            duration = rule.get('duration') or {}
            if not isinstance(matchers, list) or not isinstance(duration, dict):
                continue
            _duration_rules.append({
                'id': rule.get('id') or '',
                'name': rule.get('name') or '耗时规则',
                'priority': int(rule.get('priority') or 0),
                'match_mode': rule.get('matchMode') or rule.get('match_mode') or 'all',
                'matchers': matchers,
                'duration': duration,
            })
        _duration_rules.sort(key=lambda r: r['priority'], reverse=True)

    def _rule_field_value(rec, field):
        if field == 'folder':
            return rec.get('folder') or rec.get('folder_name') or ''
        return rec.get(field) or ''

    def _matcher_ok(rec, matcher):
        field = matcher.get('field') or 'folder'
        op = matcher.get('op') or 'contains'
        expected = str(matcher.get('value') or '').strip()
        actual = str(_rule_field_value(rec, field) or '').strip()
        if not expected:
            return False
        if op == 'equals':
            return actual.lower() == expected.lower()
        if op == 'not_contains':
            return expected.lower() not in actual.lower()
        if op == 'regex':
            try:
                return re.search(expected, actual, re.I) is not None
            except re.error:
                return expected.lower() in actual.lower()
        return expected.lower() in actual.lower()

    def _match_duration_rule(rec):
        for rule in _duration_rules:
            matchers = rule.get('matchers') or []
            if not matchers:
                continue
            results = [_matcher_ok(rec, m) for m in matchers if isinstance(m, dict)]
            if not results:
                continue
            if rule.get('match_mode') == 'any':
                if any(results):
                    return rule
            elif all(results):
                return rule
        return None

    def _duration_number(duration, key, fallback_keys=None, default=0.0):
        keys = [key] + list(fallback_keys or [])
        for k in keys:
            value = duration.get(k)
            if value is None:
                continue
            try:
                return float(value)
            except (TypeError, ValueError):
                continue
        return float(default)

    def _duration_range_value(duration, low_key, high_key, fallback_keys=None, variant='normal'):
        fallback_keys = list(fallback_keys or [])
        low = duration.get(low_key)
        high = duration.get(high_key)
        if low is None and high is None:
            base = _duration_number(duration, fallback_keys[0], fallback_keys[1:], 0.0) if fallback_keys else 0.0
            low = base
            high = base
        elif low is None:
            low = high
        elif high is None:
            high = low
        try:
            low = float(low)
        except (TypeError, ValueError):
            low = 0.0
        try:
            high = float(high)
        except (TypeError, ValueError):
            high = low
        if high < low:
            low, high = high, low
        return low if variant == 'compressed' else high

    _rule_floor_warning_keys = set()

    def _warn_rule_piece_floor(rule, raw_minutes):
        try:
            raw = float(raw_minutes)
        except (TypeError, ValueError):
            return
        if raw <= 0 or raw >= MIN_TPP:
            return
        key = rule.get('id') or rule.get('name') or 'duration_rule'
        if key in _rule_floor_warning_keys:
            return
        _rule_floor_warning_keys.add(key)
        schedule_warnings.append(
            f'耗时规则“{rule.get("name") or "耗时规则"}”的每件耗时低于{MIN_TPP}分钟，已自动按{MIN_TPP}分钟/件计算')

    def _duration_from_rule(rule, rec, variant='normal'):
        duration = rule.get('duration') or {}
        mode = duration.get('mode') or 'per_piece'
        qty = rec.get('quantity') if isinstance(rec.get('quantity'), int) and rec.get('quantity') > 0 else 0
        compressible = duration.get('compressible')
        effective_variant = variant if compressible is not False else 'normal'
        if mode == 'per_package':
            return _duration_range_value(
                duration, 'minMinutes', 'maxMinutes',
                ['minutes', 'packageMinutes', 'package_minutes'],
                effective_variant,
            )
        if mode in ('max_package_piece', 'package_piece'):
            package_minutes = _duration_range_value(
                duration, 'packageMinMinutes', 'packageMaxMinutes',
                ['packageMinutes', 'package_minutes', 'minutes'],
                effective_variant,
            )
            piece_minutes = _duration_range_value(
                duration, 'pieceMinMinutes', 'pieceMaxMinutes',
                ['pieceMinutes', 'piece_minutes'],
                effective_variant,
            )
            piece_total = qty * piece_minutes if qty > 0 else 0.0
            policy = duration.get('quantityPolicy') or duration.get('quantity_policy') or ('max' if mode == 'max_package_piece' else 'package_first')
            if qty > 0 and policy in ('piece_first', 'min', 'max') and 0 < piece_minutes < MIN_TPP:
                _warn_rule_piece_floor(rule, piece_minutes)
                piece_minutes = MIN_TPP
                piece_total = qty * piece_minutes
            if policy == 'piece_first':
                return piece_total if qty > 0 else package_minutes
            if policy == 'min':
                return min(package_minutes, piece_total) if qty > 0 else package_minutes
            if policy == 'max':
                return max(package_minutes, piece_total) if qty > 0 else package_minutes
            return package_minutes
        minutes = _duration_range_value(
            duration, 'minMinutes', 'maxMinutes',
            ['minutes', 'pieceMinutes', 'piece_minutes'],
            effective_variant,
        )
        if 0 < minutes < MIN_TPP:
            _warn_rule_piece_floor(rule, minutes)
            minutes = MIN_TPP
        return minutes * (qty if qty > 0 else 1)

    def _rule_has_compression(rule, rec):
        return _duration_from_rule(rule, rec, 'compressed') + 0.001 < _duration_from_rule(rule, rec, 'normal')

    def _rule_source(rule):
        name = rule.get('name') or '耗时规则'
        return f"duration_rule:{name}"

    # 预处理记录
    valid_records = []
    for r in records:
        rec = dict(r)
        if rec.get('manual_quantity') is not None:
            rec['quantity'] = rec['manual_quantity']
        has_duration = rec.get('manual_duration') is not None and rec['manual_duration'] > 0
        has_qty = isinstance(rec.get('quantity'), int) and rec['quantity'] > 0
        rec['_is_zhengxing_cnc'] = _is_zhengxing_cnc(rec.get('folder', ''))
        has_duration_rule = _match_duration_rule(rec) is not None
        has_package_duration = has_duration_rule or rec.get('station') == 'CNC' or rec.get('_is_zhengxing_cnc')
        if not has_duration and not has_qty and not has_package_duration:
            schedule_warnings.append(f"{rec['folder']}: 缺少件数和测量时间，已跳过排程")
            continue
        valid_records.append(rec)
    records = valid_records

    # 计算 natural_tpp（排除 CNC、耗时规则、真实手量、整形 CNC）
    qtys = [r['quantity'] for r in records
            if isinstance(r.get('quantity'), int) and r['quantity'] > 0
            and r.get('station') != 'CNC'
            and not r.get('_is_zhengxing_cnc')
            and _match_duration_rule(r) is None
            and _match_special(r) is None]
    min_qty = min(qtys) if qtys else 1
    max_qty = max(qtys) if qtys else 1
    same_qty = (min_qty == max_qty) or len(qtys) <= 1

    natural_tpps = []
    for r in records:
        if _match_duration_rule(r) is not None:
            natural_tpps.append(None)
        elif r.get('station') == 'CNC' and not r.get('_is_zhengxing_cnc'):
            natural_tpps.append(None)
        elif r.get('_is_zhengxing_cnc'):
            natural_tpps.append(None)
        elif _match_special(r) is not None:
            natural_tpps.append(None)
        elif isinstance(r.get('quantity'), int) and r['quantity'] > 0:
            if same_qty:
                natural_tpps.append(avg_tpp)
            else:
                ratio = ((r['quantity'] - min_qty) / (max_qty - min_qty)) ** 0.5
                natural_tpps.append(tpp_max - ratio * (tpp_max - tpp_min))
        else:
            natural_tpps.append(None)

    # 规范化真实手量任务
    _real_manual_tasks = []
    if real_manual_tasks:
        for mt in real_manual_tasks:
            if not isinstance(mt, dict):
                continue
            duration = mt.get('duration_minutes')
            try:
                duration = float(duration)
            except (TypeError, ValueError):
                continue
            if duration <= 0:
                continue
            task = dict(mt)
            task['_dur'] = duration
            task['station'] = mt.get('station') or '手量'
            task['note'] = '手量'
            task['manual_kind'] = 'real'
            # 使用 id 或生成稳定标识，用于判断同一任务的多段
            task['manual_task_id'] = str(task.get('id') or task.get('manual_task_id') or f"manual_{len(_real_manual_tasks)}")
            _real_manual_tasks.append(task)

    # 将真实手量拼接到 records 末尾参与排程
    if _real_manual_tasks:
        records = list(records) + _real_manual_tasks
        for _ in _real_manual_tasks:
            natural_tpps.append(None)

    # 如果既没有普通记录也没有真实手量，直接返回空
    if not records:
        return [], ['没有可排程的任务（普通记录和真实手量均为空）']

    regular_qty_total = sum(
        r.get('quantity') for i, r in enumerate(records)
        if i < len(natural_tpps)
        and natural_tpps[i] is not None
        and isinstance(r.get('quantity'), int)
        and r.get('quantity') > 0
    )

    def _regular_max_scale():
        ordinary_nats = [nat for nat in natural_tpps if nat is not None and nat > 0]
        if not ordinary_nats:
            return 1.0
        return max(1.0, tpp_max / min(ordinary_nats))

    def _source_total(durations, sources, wanted_source):
        return sum(float(dur) for dur, source in zip(durations, sources) if source == wanted_source)

    def build_segments(cur, total_dur, rec_info, seg_type='work', breaks_list=None):
        segments = []
        remaining = total_dur
        _breaks = breaks_list if breaks_list is not None else breaks
        while remaining > 0.001:
            placed = False
            for bs, be in _breaks:
                if cur >= be:
                    continue
                if cur < bs:
                    before = min(remaining, bs - cur)
                    if before > 0.001:
                        seg = dict(rec_info)
                        seg.update({'start': cur, 'end': cur + before, 'type': seg_type})
                        segments.append(seg)
                        remaining -= before
                        cur += before
                    if remaining <= 0.001:
                        placed = True
                        break
                if remaining > 0.001 and cur < be:
                    segments.append({'type': 'rest', 'start': cur, 'end': be})
                    cur = be
                    placed = True
                    break
            if not placed:
                if remaining > 0.001:
                    seg = dict(rec_info)
                    seg.update({'start': cur, 'end': cur + remaining, 'type': seg_type})
                    segments.append(seg)
                    cur += remaining
                remaining = 0
        return segments, cur

    def try_schedule(durations, duration_sources=None, extra_slots=None, breaks_list=None):
        cur = 0
        all_seg = []
        items = list(records)
        item_durs = list(durations)
        item_sources = list(duration_sources) if duration_sources else ['tpp'] * len(records)
        if extra_slots:
            merged = sorted(extra_slots, key=lambda x: x[0])
            result, dur_result, source_result, idx, si = [], [], [], 0, 0
            while idx < len(items) or si < len(merged):
                if si < len(merged) and (idx >= len(items) or merged[si][0] <= idx):
                    result.append(merged[si][1])
                    dur_result.append(merged[si][1].get('_dur', 30.0))
                    source_result.append(merged[si][1].get('duration_source', 'hidden_buffer'))
                    si += 1
                else:
                    result.append(items[idx])
                    dur_result.append(item_durs[idx])
                    source_result.append(item_sources[idx] if idx < len(item_sources) else 'tpp')
                    idx += 1
            items = result
            item_durs = dur_result
            item_sources = source_result

        real_idx = 0
        real_count = sum(1 for it in items
                         if (it.get('duration_source') not in ('hand_filler', 'other_filler')
                             and (it.get('note') not in ('手量', '其他事务') or it.get('manual_kind') == 'real'))
                         and not it.get('hidden')
                         and it.get('type') != 'hidden_buffer')
        for j, rec in enumerate(items):
            is_filler = (
                rec.get('duration_source') in ('hand_filler', 'other_filler')
                or rec.get('note') in ('手量', '其他事务')
            ) and rec.get('manual_kind') != 'real'
            is_real_manual = rec.get('manual_kind') == 'real'
            is_hidden = rec.get('hidden') == True or rec.get('type') == 'hidden_buffer'
            dur = rec['_dur'] if is_filler or is_hidden or is_real_manual else item_durs[j]
            rec_info = {k: v for k, v in rec.items() if k != '_dur'}
            seg_type = 'hidden_buffer' if is_hidden else 'work'
            # 把 duration_source 写入 segment，方便 preview 识别耗时来源
            if not is_hidden and not is_filler and not is_real_manual and j < len(item_sources):
                rec_info['duration_source'] = item_sources[j]
            segs, cur = build_segments(cur, dur, rec_info, seg_type=seg_type, breaks_list=breaks_list)
            all_seg.extend(segs)
            if pkg_rest > 0 and not is_filler and not is_hidden and not is_real_manual:
                real_idx += 1
                if real_idx < real_count:
                    rest_segs, cur = build_segments(cur, float(pkg_rest),
                                                     {'type': 'pkg_rest'}, seg_type='pkg_rest',
                                                     breaks_list=breaks_list)
                    all_seg.extend(rest_segs)
        return all_seg, cur

    def make_durations(scale, min_tpp_override=None, rule_variant='normal'):
        durations = []
        duration_sources = []
        min_tpp_for_scale = min_tpp_override if min_tpp_override is not None else tpp_min
        for i, r in enumerate(records):
            if r.get('manual_kind') == 'real':
                durations.append(float(r['_dur']))
                duration_sources.append('real_manual')
            elif _match_duration_rule(r) is not None:
                rule = _match_duration_rule(r)
                dur = _duration_from_rule(rule, r, rule_variant)
                if dur > 0:
                    durations.append(dur)
                    source = _rule_source(rule)
                    if rule_variant == 'compressed' and _rule_has_compression(rule, r):
                        source += ':compressed'
                    duration_sources.append(source)
                else:
                    durations.append(1 * max(min_tpp_for_scale, avg_tpp))
                    duration_sources.append('tpp')
            elif _match_special(r) is not None:
                sp_min = _match_special(r)
                qty = r.get('quantity') if isinstance(r.get('quantity'), int) and r['quantity'] > 0 else 1
                durations.append(float(sp_min) * qty)
                duration_sources.append('special')
            elif r.get('_is_zhengxing_cnc'):
                qty = r.get('quantity') if isinstance(r.get('quantity'), int) and r['quantity'] > 0 else 0
                dur = max(30.0, qty * 5.0) if qty > 0 else 30.0
                durations.append(dur)
                duration_sources.append('zhengxing_cnc')
            elif r.get('station') == 'CNC':
                durations.append(20.0)
                duration_sources.append('cnc')
            elif r.get('manual_duration') is not None and r['manual_duration'] > 0:
                durations.append(float(r['manual_duration']))
                duration_sources.append('manual_duration')
            else:
                nat = natural_tpps[i]
                tpp = max(min_tpp_for_scale, min(tpp_max, (nat or avg_tpp) * scale))
                if isinstance(r.get('quantity'), int) and r['quantity'] > 0:
                    durations.append(r['quantity'] * tpp)
                else:
                    durations.append(1 * tpp)
                duration_sources.append('tpp')
        return durations, duration_sources

    def compute_detailed_stats(segs):
        """计算各类任务对有效时长的贡献，用于预览缺口诊断。"""
        stats = {
            'regular_effective': 0.0,
            'real_manual_effective': 0.0,
            'special_effective': 0.0,
            'zhengxing_cnc_effective': 0.0,
            'cnc_effective': 0.0,
            'hand_filler_minutes': 0.0,
            'other_filler_minutes': 0.0,
            'hidden_buffer_total': 0.0,
            'total_effective': 0.0,
            'total_rest': 0.0,
            'total_shift': 0.0,
            'last_end': 0.0,
        }
        for seg in segs:
            dur = seg.get('end', 0) - seg.get('start', 0)
            st = seg.get('type')
            source = seg.get('duration_source')
            note = seg.get('note')
            manual_kind = seg.get('manual_kind')
            if st == 'rest' or st == 'pkg_rest':
                stats['total_rest'] += dur
            elif st == 'hidden_buffer':
                stats['hidden_buffer_total'] += dur
            elif st in ('work',):
                if manual_kind == 'real':
                    stats['real_manual_effective'] += dur
                    stats['total_effective'] += dur
                elif source == 'hand_filler' or note == '手量':
                    stats['hand_filler_minutes'] += dur
                    stats['total_effective'] += dur
                elif source == 'other_filler' or note == '其他事务':
                    stats['other_filler_minutes'] += dur
                    stats['total_effective'] += dur
                elif isinstance(source, str) and source.startswith('duration_rule:'):
                    rule_name = source.split(':', 1)[1]
                    if '整形' in rule_name and 'CNC' in rule_name:
                        stats['zhengxing_cnc_effective'] += dur
                    elif 'CNC' in rule_name:
                        stats['cnc_effective'] += dur
                    else:
                        stats['special_effective'] += dur
                    stats['total_effective'] += dur
                elif source == 'special':
                    stats['special_effective'] += dur
                    stats['total_effective'] += dur
                elif source == 'zhengxing_cnc':
                    stats['zhengxing_cnc_effective'] += dur
                    stats['total_effective'] += dur
                elif source == 'cnc':
                    stats['cnc_effective'] += dur
                    stats['total_effective'] += dur
                else:
                    stats['regular_effective'] += dur
                    stats['total_effective'] += dur
            if seg.get('end', 0) > stats['last_end']:
                stats['last_end'] = seg.get('end', 0)
        stats['total_shift'] = stats['last_end']
        return stats

    def compute_stats(segs):
        """计算有效时长、休息时长、总跨度、hidden_buffer 总时长。"""
        stats = compute_detailed_stats(segs)
        return (stats['total_effective'], stats['total_rest'], stats['hidden_buffer_total'],
                stats['total_shift'], stats['last_end'])

    def build_time_anomaly(total_effective, target_work, last_end, target_end, durations, sources):
        """生成预览页时间异常建议；只展示建议，不修改 records。"""
        ordinary_cap = 6.0

        def _folder(rec):
            return rec.get('folder') or rec.get('folder_name') or rec.get('product') or '/'

        def _is_adjustable_ordinary(index, rec, source):
            if source != 'tpp':
                return False
            if index >= len(natural_tpps) or natural_tpps[index] is None:
                return False
            if rec.get('manual_duration') is not None and rec.get('manual_duration') > 0:
                return False
            qty = rec.get('quantity')
            return isinstance(qty, int) and qty > 0

        if last_end > target_end + 0.5:
            overrun = max(0.0, last_end - target_end)
            omit_candidates = []
            for i, rec in enumerate(records):
                source = sources[i] if i < len(sources) else ''
                if not _is_adjustable_ordinary(i, rec, source):
                    continue
                qty = rec.get('quantity')
                dur = float(durations[i]) if i < len(durations) else 0.0
                if dur <= 0:
                    continue
                omit_candidates.append({
                    'folder': _folder(rec),
                    'quantity': qty,
                    'current_per_piece': dur / qty if qty else None,
                    'current_total_minutes': dur,
                    'reason': '普通料候选；本轮只建议暂不写入，不移动文件夹。',
                })
            omit_candidates.sort(key=lambda item: item['current_total_minutes'], reverse=True)
            omit_items = []
            covered = 0.0
            for item in omit_candidates:
                omit_items.append(item)
                covered += item['current_total_minutes']
                if covered >= overrun and len(omit_items) >= 1:
                    break
                if len(omit_items) >= 5:
                    break
            return {
                'kind': 'too_much',
                'title': '任务量过多',
                'message': f'当前排程会超过目标结束时间约{int(round(overrun))}分钟，建议先核对是否有包应暂不写入本日报。',
                'overrun_minutes': overrun,
                'current_effective': total_effective,
                'min_effective': target_work,
                'omit_items': omit_items,
                'note': '未确认前不会移动、重命名或修改任何文件夹；点击确认并通过二次确认后，才会把上述候选包移入新建文件夹A/B。',
            }

        if total_effective + 0.01 >= target_work:
            return None

        shortage = max(0.0, target_work - total_effective)
        candidates = []
        for i, rec in enumerate(records):
            source = sources[i] if i < len(sources) else ''
            if not _is_adjustable_ordinary(i, rec, source):
                continue
            qty = rec.get('quantity')
            dur = float(durations[i]) if i < len(durations) else 0.0
            if qty <= 0 or dur <= 0:
                continue
            current_pp = dur / qty
            add_room = max(0.0, qty * (ordinary_cap - current_pp))
            if add_room <= 0.5:
                continue
            candidates.append({
                'folder': _folder(rec),
                'quantity': qty,
                'current_per_piece': current_pp,
                'current_total_minutes': dur,
                'add_room': add_room,
            })
        candidates.sort(key=lambda item: (item['current_per_piece'], item['quantity'], item['folder']))

        adjustment_items = []
        remaining = shortage
        for item in candidates:
            add_minutes = min(item['add_room'], remaining)
            if add_minutes <= 0.5:
                continue
            recommended_pp = min(ordinary_cap, item['current_per_piece'] + add_minutes / item['quantity'])
            recommended_total = recommended_pp * item['quantity']
            adjustment_items.append({
                'folder': item['folder'],
                'quantity': item['quantity'],
                'current_per_piece': item['current_per_piece'],
                'recommended_per_piece': recommended_pp,
                'current_total_minutes': item['current_total_minutes'],
                'recommended_total_minutes': recommended_total,
                'add_minutes': max(0.0, recommended_total - item['current_total_minutes']),
                'reason': '普通料候选；特殊规则、CNC、真实手量和已有手动覆盖不纳入推荐。',
            })
            remaining -= add_minutes
            if remaining <= 0.5:
                break
            if len(adjustment_items) >= 5:
                break

        supplemental = max(0.0, remaining)
        if adjustment_items:
            message = f'当前有效工时还缺约{int(round(shortage))}分钟，可优先把少量普通包上调；普通料建议不超过每颗6分钟。'
        else:
            message = f'当前有效工时还缺约{int(round(shortage))}分钟；没有可上调的普通包，建议补充真实手量或其他事务。'
        return {
            'kind': 'too_little',
            'title': '测料太少',
            'message': message,
            'shortage_minutes': shortage,
            'current_effective': total_effective,
            'min_effective': target_work,
            'adjustment_items': adjustment_items,
            'supplemental_minutes': supplemental,
            'note': '推荐只用于核对，不会自动保存到本日包耗时覆盖。',
        }

    def schedule_with_strategy(use_early):
        """使用指定下班策略排程，返回 (segments, end_time, breaks_used, max_shift_used)。"""
        _breaks = cfg['early_breaks'] if use_early else cfg['breaks']
        _max_shift = cfg['early_max'] if use_early else MAX_SHIFT
        _target_work = TARGET_WORK_EARLY if use_early else TARGET_WORK_NORMAL
        # 目标有效时长不含固定休息；目标下班时刻按真实钟表跨度：
        # 下早班 8 小时有效 + 休息 = 班次开始后 560 分钟；
        # 正常班 10 小时有效 + 休息 = 班次开始后 720 分钟。
        _target_end = _max_shift
        _ideal_min_end = _target_end - 10

        base_durations, base_sources = make_durations(1.0)
        segs, cur = try_schedule(base_durations, duration_sources=base_sources, breaks_list=_breaks)
        chosen_durations, chosen_sources = base_durations, base_sources
        chosen_scale = 1.0
        chosen_min_override = None
        chosen_rule_variant = 'normal'
        if cur > _target_end:
            # 任务过多时先压缩可压缩的耗时规则，再压缩普通每件耗时；
            # make_durations 会保证普通 tpp >= MIN_TPP。
            compressed_rule_durations, compressed_rule_sources = make_durations(1.0, rule_variant='compressed')
            compressed_rule_segs, compressed_rule_cur = try_schedule(
                compressed_rule_durations,
                duration_sources=compressed_rule_sources,
                breaks_list=_breaks,
            )
            if compressed_rule_cur < cur:
                schedule_warnings.append('任务量偏多，已将可压缩耗时规则按区间下限计算')
                segs, cur = compressed_rule_segs, compressed_rule_cur
                chosen_durations, chosen_sources = compressed_rule_durations, compressed_rule_sources
                chosen_rule_variant = 'compressed'

            lo, hi = 0.01, 1.0
            best_fit = None
            best_fit_cur = None
            best_fit_durations = None
            best_fit_sources = None
            best_fit_scale = chosen_scale
            min_segs, min_cur = segs, cur
            min_durations, min_sources = chosen_durations, chosen_sources
            min_scale = chosen_scale
            for _ in range(40):
                mid = (lo + hi) / 2
                trial_durations, trial_sources = make_durations(mid, min_tpp_override=MIN_TPP, rule_variant='compressed')
                trial_segs, trial_cur = try_schedule(trial_durations, duration_sources=trial_sources, breaks_list=_breaks)
                min_segs, min_cur = trial_segs, trial_cur
                min_durations, min_sources = trial_durations, trial_sources
                min_scale = mid
                if trial_cur <= _target_end:
                    best_fit, best_fit_cur = trial_segs, trial_cur
                    best_fit_durations, best_fit_sources = trial_durations, trial_sources
                    best_fit_scale = mid
                    lo = mid
                else:
                    hi = mid
            if best_fit is not None:
                segs, cur = best_fit, best_fit_cur
                chosen_durations, chosen_sources = best_fit_durations, best_fit_sources
                chosen_scale = best_fit_scale
                chosen_min_override = MIN_TPP
                chosen_rule_variant = 'compressed'
            else:
                segs, cur = min_segs, min_cur
                chosen_durations, chosen_sources = min_durations, min_sources
                chosen_scale = min_scale
                chosen_min_override = MIN_TPP
                chosen_rule_variant = 'compressed'

        current_effective, _, _, _, _ = compute_stats(segs)
        max_scale = _regular_max_scale()
        if current_effective + 0.01 < _target_work and max_scale > chosen_scale + 0.001:
            base_regular = _source_total(chosen_durations, chosen_sources, 'tpp')
            lo, hi = chosen_scale, max_scale
            best_lift = None
            best_lift_cur = None
            best_lift_durations = None
            best_lift_sources = None
            best_lift_effective = None
            for _ in range(40):
                mid = (lo + hi) / 2
                trial_durations, trial_sources = make_durations(
                    mid,
                    min_tpp_override=chosen_min_override,
                    rule_variant=chosen_rule_variant,
                )
                trial_segs, trial_cur = try_schedule(trial_durations, duration_sources=trial_sources, breaks_list=_breaks)
                trial_effective, _, _, _, _ = compute_stats(trial_segs)
                if trial_effective + 0.01 >= _target_work:
                    if trial_cur <= _target_end:
                        best_lift = trial_segs
                        best_lift_cur = trial_cur
                        best_lift_durations = trial_durations
                        best_lift_sources = trial_sources
                        best_lift_effective = trial_effective
                    hi = mid
                else:
                    lo = mid
            if best_lift is not None:
                lifted_regular = _source_total(best_lift_durations, best_lift_sources, 'tpp')
                lifted_minutes = max(0.0, lifted_regular - base_regular)
                if lifted_minutes > 0.5:
                    schedule_warnings.append(
                        f'普通料耗时已在允许范围内上调约{int(round(lifted_minutes))}分钟，优先保证当天有效工时达标')
                segs, cur = best_lift, best_lift_cur
                chosen_durations, chosen_sources = best_lift_durations, best_lift_sources
                current_effective = best_lift_effective
            else:
                best_under = None
                best_under_cur = None
                best_under_durations = None
                best_under_sources = None
                best_under_effective = current_effective
                lo, hi = chosen_scale, max_scale
                for _ in range(40):
                    mid = (lo + hi) / 2
                    trial_durations, trial_sources = make_durations(
                        mid,
                        min_tpp_override=chosen_min_override,
                        rule_variant=chosen_rule_variant,
                    )
                    trial_segs, trial_cur = try_schedule(trial_durations, duration_sources=trial_sources, breaks_list=_breaks)
                    trial_effective, _, _, _, _ = compute_stats(trial_segs)
                    if trial_cur <= _target_end:
                        best_under = trial_segs
                        best_under_cur = trial_cur
                        best_under_durations = trial_durations
                        best_under_sources = trial_sources
                        best_under_effective = trial_effective
                        lo = mid
                    else:
                        hi = mid
                if best_under is not None and best_under_effective > current_effective + 0.5:
                    lifted_regular = _source_total(best_under_durations, best_under_sources, 'tpp')
                    lifted_minutes = max(0.0, lifted_regular - base_regular)
                    if lifted_minutes > 0.5:
                        schedule_warnings.append(
                            f'普通料耗时已在允许范围内上调约{int(round(lifted_minutes))}分钟，优先减少当天有效工时缺口')
                    segs, cur = best_under, best_under_cur
                    chosen_durations, chosen_sources = best_under_durations, best_under_sources
                    current_effective = best_under_effective

        regular_upper_durations, regular_upper_sources = make_durations(
            max_scale,
            min_tpp_override=chosen_min_override,
            rule_variant=chosen_rule_variant,
        )
        regular_upper_effective = _source_total(regular_upper_durations, regular_upper_sources, 'tpp')

        return (
            segs, cur, _breaks, _max_shift, _target_work, _target_end,
            chosen_durations, chosen_sources, regular_upper_effective
        )

    def add_fillers(segs, cur, max_shift, target_work, breaks_list, base_durations, base_sources):
        """如果仍不足 target_work，插入手量/其他事务。返回新 segs, cur。"""
        effective, _, _, _, _ = compute_stats(segs)
        if effective + 0.01 >= target_work:
            return segs, cur
        need = target_work - effective
        slots = []
        slack = max(0.0, max_shift - cur)
        if enable_hand:
            hand_count = 0
            while need > 5 and hand_count < 2:
                fill = min(need, hand_max, slack)
                if fill < 5:
                    break
                if fill == slack and slack + 0.01 < need and hand_max > slack:
                    schedule_warnings.append(
                        f'手量最大时长({int(hand_max)}分钟)超过班次剩余时间({int(slack)}分钟)，已自动限制')
                slots.append(('手量', fill))
                need -= fill
                slack -= fill
                hand_count += 1
        if enable_other and need > 5:
            fill = min(need, other_max, slack)
            if fill >= 5:
                if fill == slack and slack + 0.01 < need and other_max > slack:
                    schedule_warnings.append(
                        f'其他事务最大时长({int(other_max)}分钟)超过班次剩余时间({int(slack)}分钟)，已自动限制')
                slots.append((other_note, fill, 'other_filler'))
                need -= fill
                slack -= fill

        if slots:
            n_items = len(records)
            if filler_position == 'head':
                positions = [0 for _ in slots]
            elif filler_position == 'tail':
                positions = [n_items for _ in slots]
            elif filler_position == 'random':
                seed_text = '|'.join(str(r.get('folder') or r.get('product') or '') for r in records)
                seed_text += '|' + '|'.join(f'{slot[0]}:{round(float(slot[1]), 3)}' for slot in slots)
                rng = random.Random(seed_text)
                positions = sorted(rng.randint(0, n_items) for _ in slots)
            else:
                gap = max(1, n_items // (len(slots) + 1))
                positions = [min((i + 1) * gap, n_items) for i in range(len(slots))]
            position_labels = {'head': '头部', 'middle': '中部', 'tail': '尾部', 'random': '随机'}
            schedule_warnings.append(
                f'补时间手量/其他事务已按“{position_labels.get(filler_position, "中部")}”插入；可在单日设置调整为头部/中部/尾部/随机')
            extra_slots = []
            for i, slot in enumerate(slots):
                if len(slot) == 3:
                    note, dur, duration_source = slot
                else:
                    note, dur = slot
                    duration_source = 'hand_filler' if note == '手量' else 'other_filler'
                pos = positions[i]
                extra_slots.append((pos, {
                    'station': '/', 'product': '/', 'sender': '/', 'work_order': '/',
                    'mold': '/', 'machine': '/', 'test_type': '/',
                    'send_date': '/', 'send_time': '/', 'quantity': '/',
                    'note': note, '_dur': float(dur), 'duration_source': duration_source,
                }))
            segs, cur = try_schedule(base_durations, duration_sources=base_sources, extra_slots=extra_slots, breaks_list=breaks_list)
        return segs, cur

    def append_hidden_buffer(segs, cur, target_end, breaks_list):
        """用隐形缓冲填充剩余钟表空档；不计入有效工时。"""
        ideal_min_end = target_end - 10
        if cur >= ideal_min_end:
            return segs, cur
        result = list(segs)
        while cur < ideal_min_end - 0.001:
            dur = min(10.0, ideal_min_end - cur)
            hidden_segs, cur = build_segments(cur, dur, {
                'station': '/', 'product': '/', 'sender': '/', 'work_order': '/',
                'mold': '/', 'machine': '/', 'test_type': '/',
                'send_date': '/', 'send_time': '/', 'quantity': '/',
                'type': 'hidden_buffer', 'hidden': True, 'note': '隐形缓冲',
            }, seg_type='hidden_buffer', breaks_list=breaks_list)
            result.extend(hidden_segs)
        return result, cur

    # 执行排程
    best_segs = None
    best_end = 0
    best_breaks = None
    best_max_shift = 0
    best_target_work = TARGET_WORK_NORMAL
    best_target_end = TARGET_WORK_NORMAL
    chosen_strategy = leave_strategy
    strategy_reason = ''
    best_durations = []
    best_sources = []
    best_regular_upper_effective = 0.0

    if leave_strategy == 'early':
        (best_segs, best_end, best_breaks, best_max_shift, best_target_work, best_target_end,
         best_durations, best_sources, best_regular_upper_effective) = schedule_with_strategy(True)
        if best_end > best_target_end:
            schedule_warnings.append(
                f'任务量过多，即使按每件{MIN_TPP}分钟压缩仍会超过目标结束时间，建议将部分包移到下一天')
    elif leave_strategy == 'normal':
        (best_segs, best_end, best_breaks, best_max_shift, best_target_work, best_target_end,
         best_durations, best_sources, best_regular_upper_effective) = schedule_with_strategy(False)
        if best_end > best_target_end:
            schedule_warnings.append(
                f'任务量过多，即使按每件{MIN_TPP}分钟压缩仍会超过目标结束时间，建议将部分包移到下一天')
    else:  # auto
        (segs_normal, end_normal, breaks_normal, max_normal, target_work_normal, target_end_normal,
         durations_normal, sources_normal, regular_upper_normal) = schedule_with_strategy(False)
        best_segs, best_end, best_breaks, best_max_shift, best_target_work, best_target_end = \
            segs_normal, end_normal, breaks_normal, max_normal, target_work_normal, target_end_normal
        best_durations, best_sources, best_regular_upper_effective = durations_normal, sources_normal, regular_upper_normal
        chosen_strategy = 'normal'
        if end_normal <= target_end_normal:
            strategy_reason = '默认按正常班排程；如需下早班请显式选择下早班'
        else:
            strategy_reason = '正常班仍超过目标下班时间'
            schedule_warnings.append(
                f'任务量过多，即使按每件{MIN_TPP}分钟压缩仍会超过目标结束时间，建议将部分包移到下一天')

    breaks = best_breaks
    max_shift = best_max_shift
    target_work = best_target_work
    target_end = best_target_end

    # 插入手量/其他事务补时
    best_segs, best_end = add_fillers(best_segs, best_end, max_shift, target_work, breaks, best_durations, best_sources)
    # 最后再用隐形缓冲填充剩余空档。缓冲只用于时间轴显示，不参与有效工时判断。
    best_segs, best_end = append_hidden_buffer(best_segs, best_end, target_end, breaks)

    # 最终检查
    detailed_stats = compute_detailed_stats(best_segs)
    total_effective = detailed_stats['total_effective']
    total_rest = detailed_stats['total_rest']
    hidden_buffer_total = detailed_stats['hidden_buffer_total']
    total_shift = detailed_stats['total_shift']
    last_end = detailed_stats['last_end']
    time_anomaly = build_time_anomaly(
        total_effective,
        target_work,
        last_end,
        target_end,
        best_durations,
        best_sources,
    )
    if total_effective + 0.01 < target_work:
        duration_str = f'{int(total_effective // 60)}h{int(total_effective % 60)}m'
        target_str = f'{int(target_work // 60)}h{int(target_work % 60)}m'
        shortage_minutes = max(0.0, target_work - total_effective)
        if enable_hand or enable_other:
            schedule_warnings.append(
                f'有效时长不足: 当前{duration_str}, 需要{target_str}, 即使插入手量/其他事务仍不够')
            inserted_parts = []
            if enable_hand:
                inserted_parts.append(
                    f'手量已插入约{int(round(detailed_stats["hand_filler_minutes"]))}分钟，上限{int(round(hand_max))}分钟/段')
            if enable_other:
                inserted_parts.append(
                    f'其他事务已插入约{int(round(detailed_stats["other_filler_minutes"]))}分钟，上限{int(round(other_max))}分钟/段')
            hidden_text = (
                f'；当前还有约{int(round(hidden_buffer_total))}分钟隐形缓冲，但隐形缓冲不计入有效工时'
                if hidden_buffer_total > 0.5 else ''
            )
            schedule_warnings.append(
                f'补时仍不足: 受当前补时上限或可插入空档限制，{"、".join(inserted_parts)}，仍差约{int(round(shortage_minutes))}分钟{hidden_text}。可提高补时上限、增加可计入事务，或显式选择下早班。')
        else:
            schedule_warnings.append(
                f'有效时长不足: 当前{duration_str}, 需要{target_str}, 请勾选手量或其他事务补时')

    # 在第一个 segment 中记录排程元信息，供 preview 使用
    if best_segs:
        best_segs[0]['_schedule_meta'] = {
            'strategy': chosen_strategy,
            'reason': strategy_reason,
            'target_work': target_work,
            'target_end': target_end,
            'total_effective': total_effective,
            'total_rest': total_rest,
            'hidden_buffer_total': hidden_buffer_total,
            'total_shift': total_shift,
            'last_end': last_end,
            'regular_effective': detailed_stats['regular_effective'],
            'real_manual_effective': detailed_stats['real_manual_effective'],
            'special_effective': detailed_stats['special_effective'],
            'zhengxing_cnc_effective': detailed_stats['zhengxing_cnc_effective'],
            'cnc_effective': detailed_stats['cnc_effective'],
            'hand_filler_minutes': detailed_stats['hand_filler_minutes'],
            'other_filler_minutes': detailed_stats['other_filler_minutes'],
            'regular_quantity': regular_qty_total,
            'regular_avg_tpp': (detailed_stats['regular_effective'] / regular_qty_total) if regular_qty_total > 0 else 0,
            'regular_tpp_max': tpp_max,
            'regular_tpp_headroom_minutes': max(0.0, best_regular_upper_effective - detailed_stats['regular_effective']),
            'regular_tpp_at_upper': best_regular_upper_effective <= detailed_stats['regular_effective'] + 0.5,
            'time_anomaly': time_anomaly,
        }

    for seg in best_segs:
        for key in ('start', 'end'):
            if key in seg and isinstance(seg[key], (int, float)):
                total = int(seg[key] + offset)
                seg[key] = dtime((total // 60) % 24, total % 60)
    return best_segs, schedule_warnings


def unique_report_path(out_dir, base_name):
    first_path = os.path.join(out_dir, f"{base_name}.xlsx")
    if not os.path.exists(first_path):
        return first_path
    for index in range(1, 1000):
        candidate = os.path.join(out_dir, f"{base_name}（{index}）.xlsx")
        if not os.path.exists(candidate):
            return candidate
    timestamp = datetime.now().strftime("_%H%M%S")
    return os.path.join(out_dir, f"{base_name}{timestamp}.xlsx")


def infer_year_from_path(path):
    parts = re.split(r'[\\/]+', os.path.normpath(path))
    for part in reversed(parts):
        for match in re.finditer(r'(20\d{2})', part):
            year = int(match.group(1))
            if 2020 <= year <= 2100:
                return year
    return datetime.now().year


def infer_report_date(base_dir):
    folder_name = os.path.basename(base_dir.rstrip('\\/'))
    match = re.match(r'(\d+)\.(\d+)', folder_name)
    if not match:
        return datetime.now()
    year = infer_year_from_path(base_dir)
    try:
        return datetime(year, int(match.group(1)), int(match.group(2)))
    except ValueError:
        return datetime.now()


def generate_report(records, tasks, test_date, output_name_suffix="", operator_name=OPERATOR_NAME, output_dir=None):
    base_name = f"滁州量测室总体日报汇总表-OMM-{operator_name}{output_name_suffix}"
    out_dir = output_dir if output_dir else OUTPUT_DIR
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    output_path = unique_report_path(out_dir, base_name)

    if not TEMPLATE_PATH or not os.path.isfile(TEMPLATE_PATH):
        print("错误：未找到模板文件。请在工作目录中放置模板：数据源备份/滁州量测室总体日报汇总表-OMM-禹欣1.xlsx")
        return None

    shutil.copy(TEMPLATE_PATH, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    ws.title = output_name_suffix.lstrip('-')
    for name in wb.sheetnames:
        if name != ws.title:
            del wb[name]

    # 隐形缓冲/包间休息不写入 Excel。包间休息只影响排程时间，不作为报表行展示。
    visible_tasks = [
        t for t in tasks
        if not (t.get('hidden') == True or t.get('type') in ('invisible_rest', 'hidden_buffer', 'pkg_rest'))
    ]

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 8 and mr.max_row <= 9:
            ws.unmerge_cells(str(mr))

    black_font = InlineFont(rFont='微软雅黑', sz=11, b=True, color='FF000000')
    red_font = InlineFont(rFont='微软雅黑', sz=11, b=True, color='FFFF0000')
    data_font = Font(name='微软雅黑', size=11)
    thin_side = Side(style='thin', color='FF000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws['H2'].value = CellRichText([TextBlock(text="送测项目\n", font=black_font),
                                   TextBlock(text="(OMM/CMM/3D轮廓仪）", font=red_font)])
    ws['K2'].value = "测试数量"

    for row in range(3, 41):
        for col in range(1, 19):
            ws.cell(row=row, column=col).value = None

    for row in range(3, 41):
        ws.cell(row=row, column=15).value = f'=IF(N{row}>=M{row},N{row}-M{row},IF(N{row}+1-M{row}<=0.5,N{row}+1-M{row},"检查"))'
        ws.cell(row=row, column=16).value = f'=IF(ISBLANK(M{row}),"待测","已完成")'
    for row in range(3 + len(visible_tasks), 41):
        ws.cell(row=row, column=15).value = f"=N{row}-M{row}"

    ROW_END = 3 + len(visible_tasks)
    for i, task in enumerate(visible_tasks):
        row = 3 + i
        is_rest = task.get('type') in ('rest', 'pkg_rest')
        is_fake = task.get('duration_source') in ('hand_filler', 'other_filler') or task.get('note') in ('手量', '其他事务')

        if is_rest:
            for col in range(1, 12):
                ws.cell(row=row, column=col).value = '/'
            ws.cell(row=row, column=12).value = test_date
            ws.cell(row=row, column=12).number_format = 'yyyy/m/d'
            ws.cell(row=row, column=13).value = task.get('start', '/')
            ws.cell(row=row, column=13).number_format = 'h:mm'
            ws.cell(row=row, column=14).value = task.get('end', '/')
            ws.cell(row=row, column=14).number_format = 'h:mm'
            ws.cell(row=row, column=17).value = '/'
            ws.cell(row=row, column=18).value = '休息'
            continue
        is_real_manual = task.get('manual_kind') == 'real'
        # 判断是否为同一条真实手量的后续段（非第一段）
        is_real_manual_continuation = False
        if is_real_manual:
            task_id = task.get('manual_task_id')
            if task_id:
                # 统计该任务在 visible_tasks 中已出现次数
                seen_count = sum(1 for prev in visible_tasks[:i] if prev.get('manual_task_id') == task_id)
                is_real_manual_continuation = seen_count > 0

        if is_fake and not is_real_manual:
            for col in (1, 2, 3, 4, 5, 6, 7, 10, 11):
                ws.cell(row=row, column=col).value = '/'
            ws.cell(row=row, column=8).value = "OMM"
            ws.cell(row=row, column=9).value = test_date
            ws.cell(row=row, column=9).number_format = 'yyyy/m/d'
            ws.cell(row=row, column=12).value = test_date
            ws.cell(row=row, column=12).number_format = 'yyyy/m/d'
        else:
            ws.cell(row=row, column=1).value = task.get('station', '/')
            if is_real_manual and is_real_manual_continuation:
                # 后续续行：避免重复显示完整数量，避免看起来像测了两次
                ws.cell(row=row, column=2).value = task.get('product', '/')  # 保留品名便于识别
                ws.cell(row=row, column=3).value = '/'
                ws.cell(row=row, column=4).value = '/'
                ws.cell(row=row, column=5).value = '/'
                ws.cell(row=row, column=6).value = '/'
                ws.cell(row=row, column=7).value = '/'
                ws.cell(row=row, column=8).value = "OMM"
                ws.cell(row=row, column=9).value = '/'
                ws.cell(row=row, column=10).value = '/'
                ws.cell(row=row, column=11).value = '/'  # 数量不写完整值
            else:
                ws.cell(row=row, column=2).value = task.get('product', '/')
                ws.cell(row=row, column=3).value = task.get('sender', '/')
                ws.cell(row=row, column=4).value = task.get('work_order', '/')
                ws.cell(row=row, column=5).value = task.get('mold', '/')
                ws.cell(row=row, column=6).value = task.get('machine', '/')
                ws.cell(row=row, column=7).value = task.get('test_type', '/')
                ws.cell(row=row, column=8).value = "OMM"
                ws.cell(row=row, column=9).value = task.get('send_date', '/')
                ws.cell(row=row, column=9).number_format = 'yyyy/m/d'
                send_time = _parse_time_str(task.get('send_time', '/'))
                ws.cell(row=row, column=10).value = send_time if send_time else task.get('send_time', '/')
                ws.cell(row=row, column=10).number_format = 'h:mm'
                ws.cell(row=row, column=11).value = task.get('quantity', '/')
            ws.cell(row=row, column=12).value = test_date
            ws.cell(row=row, column=12).number_format = 'yyyy/m/d'

        ws.cell(row=row, column=13).value = task.get('start', '/')
        ws.cell(row=row, column=13).number_format = 'h:mm'
        ws.cell(row=row, column=14).value = task.get('end', '/')
        ws.cell(row=row, column=14).number_format = 'h:mm'
        ws.cell(row=row, column=17).value = task.get('operator', operator_name)
        if is_real_manual:
            ws.cell(row=row, column=18).value = '手量'
        else:
            ws.cell(row=row, column=18).value = task.get('note', '')

    # 应用数据行基础格式，并清除原条件格式（原模板条件格式在删改行列后容易异常）
    for row in range(3, 41):
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            if cell.font is None or cell.font.name != '微软雅黑':
                cell.font = data_font
            cell.border = thin_border
            if col <= 16:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')
        ws.cell(row=row, column=9).number_format = 'yyyy/m/d'
        ws.cell(row=row, column=10).number_format = 'h:mm'
        ws.cell(row=row, column=12).number_format = 'yyyy/m/d'
        ws.cell(row=row, column=13).number_format = 'h:mm'
        ws.cell(row=row, column=14).number_format = 'h:mm'
        ws.cell(row=row, column=15).number_format = 'h:mm:ss;@'
    ws.conditional_formatting._cf_rules.clear()

    # 重新添加同事模板口径的条件格式：时间/耗时/量测员为绿色字体，状态列已完成/待测红绿提示。
    # 状态列底色来自条件格式，因此 Excel 工具栏可能仍显示“无填充”，这是正常的。
    last_data_row = 40
    data_green_font = Font(name='微软雅黑', size=11, color='FF006102')
    for row in range(3, last_data_row + 1):
        for col in (12, 13, 14, 15, 17):
            ws.cell(row=row, column=col).font = data_green_font

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='FF006100', bold=True)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='FF9C0006', bold=True)
    status_range = f'P3:P{last_data_row}'
    ws.conditional_formatting.add(status_range,
                                  FormulaRule(formula=['NOT(ISERROR(SEARCH("已完成",P3)))'],
                                              fill=green_fill, font=green_font))
    ws.conditional_formatting.add(status_range,
                                  FormulaRule(formula=['NOT(ISERROR(SEARCH("待测",P3)))'],
                                              fill=red_fill, font=red_font))

    wb.save(output_path)
    print(f"报表已生成: {output_path}")
    print(f"共填入 {len(tasks)} 行数据")
    return output_path


def run(base_dir, early_leave=False, leave_strategy=None,
        tpp_min=TPP_MIN_DEFAULT, tpp_max=TPP_MAX_DEFAULT,
        pkg_rest=PKG_REST_DEFAULT,
        enable_hand=True, enable_other=False,
        operator_name=OPERATOR_NAME, output_dir=None, shift_override=None,
        special_items=None, duration_rules=None, hand_max=120, other_max=90,
        real_manual_tasks=None, filler_position='middle',
        other_note='其他事务', known_senders=None, measurement_people=None):
    records, review_map = parse_all_folders(
        base_dir,
        operator_name=operator_name,
        known_senders=known_senders,
        measurement_people=measurement_people,
    )
    # 允许普通 records 为空，只要 real_manual_tasks 非空
    if not records and not real_manual_tasks:
        return None, [], [], [], {}

    folder_name = os.path.basename(base_dir.rstrip('\\/'))
    if shift_override in ('A', 'B'):
        shift_label = shift_override
    else:
        shift_label = 'B' if folder_name.upper().endswith('B') else 'A'

    segments, sched_warnings = schedule_tasks(records, shift_label, early_leave=early_leave,
                              leave_strategy=leave_strategy,
                              tpp_min=tpp_min, tpp_max=tpp_max, pkg_rest=pkg_rest,
                              enable_hand=enable_hand, enable_other=enable_other,
                              special_items=special_items, duration_rules=duration_rules,
                              hand_max=hand_max, other_max=other_max,
                              real_manual_tasks=real_manual_tasks,
                              other_note=other_note,
                              filler_position=filler_position)

    suffix = f"-{folder_name}"
    if leave_strategy == 'early' or (leave_strategy is None and early_leave):
        suffix += "-下早班"

    test_date = infer_report_date(base_dir)

    # 默认输出到源文件夹
    if output_dir is None:
        output_dir = base_dir
    output_path = generate_report(records, segments, test_date, suffix, operator_name=operator_name, output_dir=output_dir)
    return output_path, records, review_map, [], sched_warnings


def main():
    print(f"=== 禹欣日报生成系统 ===")
    print(f"数据源: {os.path.join(WORK_DIR, '数据源备份')}")
    print()
    base_dir = sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith('--') else os.path.join(WORK_DIR, "数据源备份", "6.5B")
    early = '--early' in sys.argv
    hand = '--hand' in sys.argv or '--手动量' in sys.argv
    other = '--other' in sys.argv or '--其他' in sys.argv
    output_path, records, review_map, warnings, sched_warnings = run(base_dir, early_leave=early, enable_hand=hand, enable_other=other)
    if warnings:
        for folder, warns in warnings:
            print(f"  ! {folder}: {', '.join(warns)}")
    if sched_warnings:
        for w in sched_warnings:
            print(f"  ! {w}")
    if output_path:
        print(f"\n完成！报表保存至: {output_path}")
        subprocess.Popen(f'explorer "{OUTPUT_DIR}"')
    else:
        print("未找到有效任务！")


def preview(base_dir, early_leave=False, leave_strategy=None,
            tpp_min=TPP_MIN_DEFAULT, tpp_max=TPP_MAX_DEFAULT,
            pkg_rest=PKG_REST_DEFAULT, enable_hand=True, enable_other=False,
            operator_name=OPERATOR_NAME, records=None, shift_override=None,
            special_items=None, duration_rules=None, hand_max=120, other_max=90,
            real_manual_tasks=None, known_senders=None, measurement_people=None, filler_position='middle',
            other_note='其他事务'):
    """返回预估信息字典，不生成文件。可传入已审核的 records，否则重新解析。"""
    import math

    if records is None:
        records, _ = parse_all_folders(
            base_dir,
            operator_name=operator_name,
            known_senders=known_senders,
            measurement_people=measurement_people,
        )
    # 允许 records 为空，只要 real_manual_tasks 非空；两者都空才失败
    if not records and not real_manual_tasks:
        return None

    folder_name = os.path.basename(base_dir.rstrip('\\/'))
    if shift_override in ('A', 'B'):
        shift_label = shift_override
    else:
        shift_label = 'B' if folder_name.upper().endswith('B') else 'A'
    cfg = SHIFTS[shift_label]
    start_offset = cfg['start_offset']

    # 兼容旧参数
    if leave_strategy is None:
        leave_strategy = 'early' if early_leave else 'normal'

    segments, sched_warnings = schedule_tasks(records, shift_label, early_leave=early_leave,
                              leave_strategy=leave_strategy,
                              tpp_min=tpp_min, tpp_max=tpp_max, pkg_rest=pkg_rest,
                              enable_hand=enable_hand, enable_other=enable_other,
                              special_items=special_items, duration_rules=duration_rules,
                              hand_max=hand_max, other_max=other_max,
                              real_manual_tasks=real_manual_tasks,
                              other_note=other_note,
                              filler_position=filler_position)

    # 从第一个 segment 的元信息读取统计
    meta = segments[0].get('_schedule_meta', {}) if segments else {}
    total_effective = meta.get('total_effective', 0)
    total_rest = meta.get('total_rest', 0)
    hidden_buffer_total = meta.get('hidden_buffer_total', 0)
    total_shift = meta.get('total_shift', 0)
    last_end_min = meta.get('last_end', 0)
    required_effective = meta.get('target_work', TARGET_WORK_NORMAL)
    target_end_min = meta.get('target_end', TARGET_WORK_NORMAL)
    chosen_strategy = meta.get('strategy', leave_strategy)

    target_clock_total = int(target_end_min + start_offset)
    target_clock_end = f"{(target_clock_total // 60) % 24:02d}:{target_clock_total % 60:02d}"
    actual_clock_total = int(last_end_min + start_offset)
    actual_last_end = f"{(actual_clock_total // 60) % 24:02d}:{actual_clock_total % 60:02d}"
    finish_delta_minutes = int(target_end_min - last_end_min)

    # 缺口诊断与决策建议
    visible_effective = max(0, total_effective)
    need_minutes = max(0, required_effective - visible_effective)
    if need_minutes < 0.01:
        need_minutes = 0
    shortage_level = 'ok'
    if need_minutes >= 180:
        shortage_level = 'extreme'
    elif need_minutes >= 60:
        shortage_level = 'severe'
    elif need_minutes > 0:
        shortage_level = 'shortage'

    def format_approx_minutes(minutes):
        return f'约 {int(round(max(0, minutes)))} 分钟'

    need_minutes_text = format_approx_minutes(need_minutes)

    decision = {
        'level': shortage_level,
        'need_minutes': need_minutes,
        'regular_effective': meta.get('regular_effective', 0),
        'real_manual_effective': meta.get('real_manual_effective', 0),
        'special_effective': meta.get('special_effective', 0),
        'zhengxing_cnc_effective': meta.get('zhengxing_cnc_effective', 0),
        'cnc_effective': meta.get('cnc_effective', 0),
        'hand_filler_minutes': meta.get('hand_filler_minutes', 0),
        'other_filler_minutes': meta.get('other_filler_minutes', 0),
        'hidden_buffer_total': hidden_buffer_total,
        'regular_quantity': meta.get('regular_quantity', 0),
        'regular_avg_tpp': meta.get('regular_avg_tpp', 0),
        'regular_tpp_max': meta.get('regular_tpp_max', tpp_max),
        'regular_tpp_headroom_minutes': meta.get('regular_tpp_headroom_minutes', 0),
        'regular_tpp_at_upper': meta.get('regular_tpp_at_upper', False),
        'target_clock_end': target_clock_end,
        'actual_last_end': actual_last_end,
        'options': [],
    }
    time_anomaly = meta.get('time_anomaly')
    if isinstance(time_anomaly, dict):
        time_anomaly['target_clock_end'] = target_clock_end
        time_anomaly['actual_last_end'] = actual_last_end
        decision['time_anomaly'] = time_anomaly

    if shortage_level != 'ok':
        if shortage_level == 'extreme':
            decision['message'] = f'数据严重不足：当前可见有效时长 {format_minutes(visible_effective)}，目标 {format_minutes(required_effective)}，还缺 {need_minutes_text}。'
            decision['title'] = '数据严重不足'
        elif shortage_level == 'severe':
            decision['message'] = f'数据不足：当前可见有效时长 {format_minutes(visible_effective)}，目标 {format_minutes(required_effective)}，还缺 {need_minutes_text}。'
            decision['title'] = '数据不足'
        else:
            decision['message'] = f'有效时长不足：当前可见有效时长 {format_minutes(visible_effective)}，目标 {format_minutes(required_effective)}，还缺 {need_minutes_text}。'
            decision['title'] = '有效时长不足'

        decision['options'].append({
            'key': 'as_is',
            'label': '如实填写',
            'description': '不自动补时间，按当前数据生成报表（隐形缓冲仍会显示但不写入Excel）。',
        })
        decision['options'].append({
            'key': 'add_real_manual',
            'label': '补充真实手量',
            'description': f'打开手量补录窗口，建议补充{need_minutes_text}的真实手量。',
        })
        if enable_hand or enable_other:
            decision['options'].append({
                'key': 'add_filler',
                'label': '补充其他事务 / 补时间手量',
                'description': f'使用现有设置插入补时间手量/其他事务（已插入 {int(meta.get("hand_filler_minutes", 0) + meta.get("other_filler_minutes", 0))} 分钟）。',
            })
        else:
            decision['options'].append({
                'key': 'enable_filler',
                'label': '启用补时间手量/其他事务',
                'description': '当前未启用补时间功能，可在单日设置中开启。',
            })
    else:
        decision['message'] = '当前有效时长已满足目标要求。'
        decision['title'] = '时长充足'

    # 若命中整形 CNC，在 warnings 中追加提示
    if meta.get('zhengxing_cnc_effective', 0) > 0:
        sched_warnings.append(f'整形 CNC 耗时规则已生效：按 max(30, 数量×5分钟) 计算耗时')

    rows = []
    seen_real_manual_ids = set()
    for i, seg in enumerate(segments):
        start = seg['start']
        end = seg['end']
        seg_type = seg['type']

        start_min = start.hour * 60 + start.minute
        end_min = end.hour * 60 + end.minute
        if end_min < start_min:
            end_min += 1440
        duration = end_min - start_min

        if seg_type == 'work':
            qty = seg.get('quantity', '/')
            if isinstance(qty, int) and qty > 0:
                tpp = round(duration / qty, 1)
            else:
                tpp = '—'
            note = seg.get('note', '')
            source = seg.get('duration_source')
            is_filler = source in ('hand_filler', 'other_filler') or note in ('手量', '其他事务')
            manual_kind = seg.get('manual_kind')
            if manual_kind == 'real':
                type_label = '真实手量'
                task_id = seg.get('manual_task_id')
                is_continuation = task_id and task_id in seen_real_manual_ids
                if task_id:
                    seen_real_manual_ids.add(task_id)
                if is_continuation:
                    qty = '/'
            elif source == 'hand_filler' or note == '手量':
                type_label = '手量'
            elif source == 'other_filler' or note == '其他事务':
                type_label = note or '其他事务'
            else:
                type_label = '工作'
            rows.append({
                'seq': i + 1,
                'product': seg.get('product', '/') if manual_kind != 'real' and not is_filler else seg.get('product', '/'),
                'qty': qty if manual_kind != 'real' and not is_filler else (qty if manual_kind == 'real' else '—'),
                'start': start.strftime('%H:%M'),
                'end': end.strftime('%H:%M'),
                'tpp': tpp if not is_filler else '—',
                'type': type_label,
                'source': 'real_manual' if manual_kind == 'real' else (
                    source if is_filler else seg.get('duration_source', 'tpp')
                ),
            })
        elif seg_type == 'rest':
            rows.append({
                'seq': i + 1,
                'product': '—',
                'qty': '—',
                'start': start.strftime('%H:%M'),
                'end': end.strftime('%H:%M'),
                'tpp': '—',
                'type': '休息'
            })
        elif seg_type == 'pkg_rest':
            rows.append({
                'seq': i + 1,
                'product': '—',
                'qty': '—',
                'start': start.strftime('%H:%M'),
                'end': end.strftime('%H:%M'),
                'tpp': '—',
                'type': '包间休息'
            })
        elif seg_type == 'hidden_buffer' or seg.get('hidden'):
            rows.append({
                'seq': i + 1,
                'product': '—',
                'qty': '—',
                'start': start.strftime('%H:%M'),
                'end': end.strftime('%H:%M'),
                'tpp': '—',
                'type': '隐形缓冲'
            })
        elif seg.get('manual_kind') == 'real' or seg_type == 'real_manual':
            rows.append({
                'seq': i + 1,
                'product': seg.get('product', '/'),
                'qty': seg.get('quantity', '/'),
                'start': start.strftime('%H:%M'),
                'end': end.strftime('%H:%M'),
                'tpp': '—',
                'type': '真实手量'
            })

    meets_required = total_effective + 0.01 >= required_effective

    # 估算还需多少件
    need = max(0, required_effective - total_effective)
    if need < 0.01:
        need = 0
    estimates = {}
    if need > 0:
        existing_qtys = [r['quantity'] for r in records
                        if isinstance(r.get('quantity'), int) and r['quantity'] > 0]
        if existing_qtys:
            avg_tpp = total_effective / sum(existing_qtys) if existing_qtys else (tpp_max + tpp_min) / 2
            conservative_tpp = tpp_max
            conservative_duration_per_piece = conservative_tpp + (pkg_rest / 10)
            conservative_pieces = math.ceil(need / conservative_duration_per_piece) if conservative_duration_per_piece > 0 else 999
            optimistic_duration_per_piece = avg_tpp + (pkg_rest / 10)
            optimistic_pieces = math.ceil(need / optimistic_duration_per_piece) if optimistic_duration_per_piece > 0 else 999
            estimates = {
                'optimistic': optimistic_pieces,
                'conservative': conservative_pieces,
                'need_minutes': need
            }
        else:
            estimates = {
                'optimistic': 999,
                'conservative': 999,
                'need_minutes': need
            }

    return {
        'folder_name': folder_name,
        'shift_label': shift_label,
        'early_leave': early_leave,
        'leave_strategy': chosen_strategy,
        'records': records,
        'warnings': [],
        'schedule_warnings': sched_warnings,
        'rows': rows,
        'summary': {
            'total_shift': total_shift,
            'total_work': total_effective,
            'total_effective': total_effective,
            'required_effective': required_effective,
            'total_rest': total_rest,
            'hidden_buffer_total': hidden_buffer_total,
            'meets_min': meets_required,
            'meets_required': meets_required,
            'estimates': estimates,
            'target_clock_end': target_clock_end,
            'actual_last_end': actual_last_end,
            'finish_delta_minutes': finish_delta_minutes,
            'regular_effective': decision['regular_effective'],
            'real_manual_effective': decision['real_manual_effective'],
            'special_effective': decision['special_effective'],
            'zhengxing_cnc_effective': decision['zhengxing_cnc_effective'],
            'cnc_effective': decision['cnc_effective'],
            'hand_filler_minutes': decision['hand_filler_minutes'],
            'other_filler_minutes': decision['other_filler_minutes'],
            'regular_quantity': decision['regular_quantity'],
            'regular_avg_tpp': decision['regular_avg_tpp'],
            'regular_tpp_max': decision['regular_tpp_max'],
            'regular_tpp_headroom_minutes': decision['regular_tpp_headroom_minutes'],
            'regular_tpp_at_upper': decision['regular_tpp_at_upper'],
            'need_minutes': decision['need_minutes'],
            'shortage_level': decision['level'],
            'time_anomaly': time_anomaly,
            'decision': decision,
        }
    }


if __name__ == '__main__':
    main()
